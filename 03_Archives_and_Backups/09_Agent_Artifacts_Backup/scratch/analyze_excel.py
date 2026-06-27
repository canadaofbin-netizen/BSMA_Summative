import openpyxl

excel_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_Actual Coding Sheet_v3.xlsx"
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

anomalies = []

# Verify header
header = [ws.cell(row=1, column=i).value for i in range(1, 48)]
if not header[0]:
    anomalies.append("Header is missing or shifted.")

# Check row by row
max_row = ws.max_row
last_id = None
ids = set()

for r in range(2, max_row + 1):
    eff_id = ws.cell(row=r, column=4).value
    if eff_id is None:
        continue # empty row
    
    # Check for duplicate IDs
    if eff_id in ids:
        anomalies.append(f"Row {r}: Duplicate Effect Size ID '{eff_id}'")
    ids.add(eff_id)
    
    inc = ws.cell(row=r, column=5).value
    if inc not in [0, 1]:
        anomalies.append(f"Row {r}: Include value is {inc}, should be 0 or 1.")
        
    # Check demographics numbers
    for col in [22, 23, 24, 25]: # Age, Age SD, Female, Tenure
        val = ws.cell(row=r, column=col).value
        if val is not None and type(val) == str and val.strip() == "999":
            anomalies.append(f"Row {r}, Col {col}: '999' is a string, should be integer.")
            
    # Check r value bounds
    r_val = ws.cell(row=r, column=45).value
    if inc == 1 and r_val is not None:
        if type(r_val) not in [int, float]:
            anomalies.append(f"Row {r}: r value '{r_val}' is type {type(r_val).__name__}, not numeric.")
        elif r_val != 999 and (r_val < -1.0 or r_val > 1.0):
            anomalies.append(f"Row {r}: r value '{r_val}' is out of bounds [-1, 1].")

print(f"Scanned {max_row} rows.")
if not anomalies:
    print("No obvious data type/bounds anomalies found in Python script.")
else:
    for a in anomalies[:20]:
        print(a)
    if len(anomalies) > 20:
        print(f"...and {len(anomalies)-20} more.")
