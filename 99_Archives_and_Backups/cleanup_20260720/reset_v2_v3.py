import os
import shutil
import openpyxl

BASE_DIR = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY"
ARCHIVE_DIR = os.path.join(BASE_DIR, "99_Archives_and_Backups", "cleanup_20260720")

# 1. Create Archive Directory
os.makedirs(ARCHIVE_DIR, exist_ok=True)
print(f"Created archive directory: {ARCHIVE_DIR}")

# 2. Archive V2 and V3
files_to_archive = ["BSMA_AI_Run_Validation2.xlsx", "BSMA_AI_Run_Validation3.xlsx"]
for f in files_to_archive:
    src = os.path.join(BASE_DIR, f)
    if os.path.exists(src):
        dst = os.path.join(ARCHIVE_DIR, f"Cancelled_Legacy_{f}")
        shutil.move(src, dst)
        print(f"Archived {f} to {dst}")
    else:
        print(f"File {f} not found, skipping archive.")

# 3. Create fresh V2 and V3 from V1
v1_path = os.path.join(BASE_DIR, "BSMA_AI_Run_Validation1.xlsx")
if not os.path.exists(v1_path):
    print("V1 source file not found!")
    exit(1)

for fresh_file in files_to_archive:
    fresh_path = os.path.join(BASE_DIR, fresh_file)
    print(f"Generating clean {fresh_file} from V1...")
    
    wb = openpyxl.load_workbook(v1_path)
    ws = wb.active
    
    # Wipe columns E (5), F (6), P (16) from row 4 to max_row
    for r in range(4, ws.max_row + 1):
        ws.cell(row=r, column=5).value = None
        ws.cell(row=r, column=6).value = None
        ws.cell(row=r, column=16).value = None
        
    wb.save(fresh_path)
    print(f"Created fresh slate: {fresh_path}")

print("Reset operation completed successfully.")
