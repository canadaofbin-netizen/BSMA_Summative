import openpyxl

excel_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_Actual Coding Sheet_v3.xlsx"
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

anomalies = []

for r in range(31, 115): # newly inserted rows
    eff_id = ws.cell(row=r, column=4).value
    if eff_id is None: continue
    
    # Check numeric columns that should strictly be numbers (except 999)
    for col, name in [(22, "Age"), (23, "Age SD"), (24, "Female"), (25, "Tenure"), (40, "Corr Mean"), (41, "Corr SD"), (42, "Corr Alpha"), (45, "r")]:
        val = ws.cell(row=r, column=col).value
        if val is not None and str(val) != "999":
            # If it's a string that looks like a range or is non-numeric, it might be an anomaly depending on rules.
            if type(val) == str:
                # Is it a string because it couldn't be parsed?
                anomalies.append(f"Row {r} (ID: {eff_id}): {name} is stored as text: '{val}'")
                
    # Check strings
    for col, name in [(8, "Author"), (26, "Occ Type"), (27, "Country")]:
        val = ws.cell(row=r, column=col).value
        if val is None or val == "" or val == "999" or val == 999:
            pass # Valid missing

if not anomalies:
    print("No formatting anomalies found in newly inserted rows!")
else:
    for a in anomalies:
        print(a)
