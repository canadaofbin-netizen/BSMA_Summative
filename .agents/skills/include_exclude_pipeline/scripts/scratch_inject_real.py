import openpyxl

data = {
    "BSMA0100": {
        "status": "0 = Exclude",
        "reason_code": "3 = Non-individual level (team/firm/org analysis)",
        "notes": "[The study conducts analysis at the team level (open source software projects) rather than the individual employee level.]. Verbatim Evidence: \"The unit of analysis was the OSS project.\""
    },
    "BSMA0101": {
        "status": "0 = Exclude",
        "reason_code": "3 = Non-individual level (team/firm/org analysis)",
        "notes": "[Team-level analysis of NPD projects]. Verbatim Evidence: \"The unit of analysis was the NPD project. OSS project information is hosted in web-based repositories which have been used repeatedly as sources of archival data for empirical studies (Mockus et al., 2002).\""
    },
    "BSMA0102": {
        "status": "0 = Exclude",
        "reason_code": "3 = Non-individual level (team/firm/org analysis)",
        "notes": "[The unit of analysis is the organization (Technology Transfer Centres), not individual employees.]. Verbatim Evidence: \"The response rate was equal to 43.92% for a total of 65 respondent TTCs (39 in the Veneto, 18 in Friuli and 8 in Trentino).\""
    },
    "BSMA0103": {
        "status": "0 = Exclude",
        "reason_code": "1 = No effect size of interest",
        "notes": "[Qualitative exploratory case study based on interviews; lacks quantitative effect sizes.]. Verbatim Evidence: \"Through the analysis of 25 semi-structured in-depth interviews carried out in two MNCs in eight countries where our focal companies have collaborations with universities and through their triangulation with secondary data, we offer a first understanding into company-based boundary spanning roles in the context of international U-I collaboration.\""
    },
    "BSMA0104": {
        "status": "0 = Exclude",
        "reason_code": "3 = Non-individual level (team/firm/org analysis)",
        "notes": "[The study analyzes data at the plant level rather than the individual employee level.]. Verbatim Evidence: \"The statistical analyses in this dissertation are based on survey data on nearly one thousand manufacturing plants in twenty-one industries where metal machining is a normal part of production processes.\""
    }
}

test_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\Test_Coding_Sheet.xlsx"
wb = openpyxl.load_workbook(test_path)
ws = wb.active

for row in range(4, ws.max_row + 1):
    art_id = str(ws.cell(row, 2).value).strip()
    if art_id in data:
        verdict = data[art_id]
        ws.cell(row=row, column=5).value = verdict["status"]
        ws.cell(row=row, column=6).value = verdict["reason_code"]
        ws.cell(row=row, column=16).value = verdict["notes"]

wb.save(test_path)
print("Real data injected successfully!")
