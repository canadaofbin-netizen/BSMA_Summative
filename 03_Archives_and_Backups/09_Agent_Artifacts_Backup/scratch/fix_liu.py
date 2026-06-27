import openpyxl
file_path = r'g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_Actual Coding Sheet_v2.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

bsb_text = "We measured individual BSBs with five items following Marrone et al.'s (2007) approach."
gpd_text = "Three items written by Lee, Pillutla, and Law (2000) were used to measure individual power distance orientation in this study."
ile_text = "We measured informal leader emergence following a social network approach and used the 1-item measure following Zhang et al.'s (2012) method."
tp_text = "Task performance was assessed through leader ratings of the target individual. Four items were used from Williams and Anderson's (1991) original scale."

for r in range(16, 23):
    ws.cell(row=r, column=32, value=bsb_text)

# Col 39 (Correlate Specific Measure)
ws.cell(row=16, column=39, value=gpd_text)
ws.cell(row=17, column=39, value=ile_text)
ws.cell(row=18, column=39, value=tp_text)
ws.cell(row=19, column=39, value=gpd_text)
ws.cell(row=20, column=39, value=tp_text)
ws.cell(row=21, column=39, value=gpd_text)
ws.cell(row=22, column=39, value=tp_text)

wb.save(file_path)
print('Updated Specific Measures for Liu 2016.')
