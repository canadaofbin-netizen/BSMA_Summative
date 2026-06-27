import re
import openpyxl

md_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\03_Shadow_Reports\Liu_2024_Shadow_Report.md"
excel_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_Actual Coding Sheet_v2.xlsx"
out_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_Actual Coding Sheet_v3.xlsx"

wb = openpyxl.load_workbook(excel_path)
ws = wb.active

start_row = 31 # Start after Zhang 2022 (row 29)

def clean(val):
    val = val.strip()
    val = val.replace('**', '').replace('*', '')
    if val == '999': return 999
    if val == '1': return 1
    if val == '0': return 0
    try: return float(val)
    except: return val

with open(md_path, 'r', encoding='utf-8') as f:
    lines = f.readlines()

in_table = False
r_idx = start_row

for line in lines:
    line = line.strip()
    if line.startswith("| Effect Size ID") or line.startswith("| **BSMA"):
        in_table = True
    if in_table and line.startswith("|") and not line.startswith("|---"):
        if "Effect Size ID" in line: continue
        cols = [c.strip() for c in line.split("|")[1:-1]]
        if not cols: continue
        
        eff_id = clean(cols[0])
        corr_var = clean(cols[1])
        # cols[2] is Correlate Specific Measure, wait Liu 2024 table has 10 columns
        # | Effect Size ID | Correlate Variable | Correlate Specific Measure | r | Alpha | Mean | SD | Min/Max | Included? | Notes |
        if len(cols) >= 9:
            corr_measure = clean(cols[2])
            r_val = clean(cols[3])
            corr_alpha = clean(cols[4])
            corr_mean = clean(cols[5])
            corr_sd = clean(cols[6])
            inc = clean(cols[8])
            notes = clean(cols[9]) if len(cols) >= 10 else ""
            
            # Map to Excel
            ws.cell(row=r_idx, column=1, value="BSMA0006") # Article ID
            ws.cell(row=r_idx, column=2, value="KY") # Coder Initials
            ws.cell(row=r_idx, column=3, value="BSMA0006.1") # Sample ID
            ws.cell(row=r_idx, column=4, value=eff_id) # Effect Size ID
            ws.cell(row=r_idx, column=5, value=1) # Include
            ws.cell(row=r_idx, column=8, value="Liu Ting") # Author
            ws.cell(row=r_idx, column=9, value=2024) # Year
            ws.cell(row=r_idx, column=12, value="1 = Journal article") # Pub Type
            ws.cell(row=r_idx, column=16, value=notes) # Notes
            ws.cell(row=r_idx, column=17, value="2 = Longitudinal") # Study Design
            ws.cell(row=r_idx, column=21, value="2 = Yes") # Int Context
            ws.cell(row=r_idx, column=22, value=33.91) # Age
            ws.cell(row=r_idx, column=23, value=6.64) # Age SD
            ws.cell(row=r_idx, column=24, value=0.5028) # Female
            ws.cell(row=r_idx, column=25, value=999) # Org Tenure
            ws.cell(row=r_idx, column=26, value="Chinese expatriates in energy") # Occ Type
            ws.cell(row=r_idx, column=27, value="Multiple") # Country
            
            # BS Measure
            ws.cell(row=r_idx, column=28, value=1) # BS Min
            ws.cell(row=r_idx, column=29, value=7) # BS Max
            ws.cell(row=r_idx, column=30, value=3) # BS Report
            ws.cell(row=r_idx, column=32, value="New scale developed by authors") # BS Specific Measure
            
            ws.cell(row=r_idx, column=36, value=corr_var) # Correlate Var
            ws.cell(row=r_idx, column=39, value=corr_measure) # Correlate Measure
            ws.cell(row=r_idx, column=40, value=corr_mean) # Correlate Mean
            ws.cell(row=r_idx, column=41, value=corr_sd) # Correlate SD
            ws.cell(row=r_idx, column=42, value=corr_alpha) # Correlate Alpha
            ws.cell(row=r_idx, column=45, value=r_val) # r
            
            r_idx += 1

wb.save(out_path)
print(f"Inserted {r_idx - start_row} rows for Liu 2024 into {out_path}.")
