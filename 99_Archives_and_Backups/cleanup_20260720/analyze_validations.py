import openpyxl
import os
from collections import Counter

FILES_TO_ANALYZE = [
    "BSMA_AI_Run_Validation3.xlsx",
    "Legacy_BSMA_AI_Run_Validation1.xlsx",
    "Legacy_BSMA_AI_Run_Validation1_revised.xlsx",
    "Legacy_BSMA_AI_Run_Validation2.xlsx",
    "Legacy_BSMA_AI_Run_Validation2_revised.xlsx"
]

def analyze_excel(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"Error reading {filepath}: {e}")
        return None
        
    total_processed = 0
    includes = 0
    excludes = 0
    exclusion_codes = Counter()
    
    for row in range(4, ws.max_row + 1):
        status = ws.cell(row, 5).value
        code = ws.cell(row, 6).value
        
        status_str = str(status).strip() if status else ""
        if status_str and status_str != "None":
            total_processed += 1
            if "1" in status_str or "Include" in status_str:
                includes += 1
            elif "0" in status_str or "Exclude" in status_str:
                excludes += 1
                
                # Parse exclusion code
                code_str = str(code).strip() if code else "Missing"
                if code_str == "None": code_str = "Missing"
                exclusion_codes[code_str] += 1
                
    return {
        "file": os.path.basename(filepath),
        "total": total_processed,
        "include": includes,
        "exclude": excludes,
        "codes": dict(exclusion_codes)
    }

print("Starting analysis...")
results = []
for f in FILES_TO_ANALYZE:
    full_path = os.path.join(r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY", f)
    if os.path.exists(full_path):
        res = analyze_excel(full_path)
        if res:
            results.append(res)
    else:
        print(f"File not found: {f}")
        
import json
print("\n--- RESULTS JSON ---")
print(json.dumps(results, indent=2))
