import openpyxl
from openpyxl.cell import MergedCell

# Create the Validation 2 worksheet by clearing out the Validation 1 decisions
# We will copy from the revised validation 1 sheet so we keep the exact formatting and latest paper list
wb = openpyxl.load_workbook('G:\\My Drive\\UCL\\BSMA\\BSMA ANTIGRAVITY\\BSMA_AI_Run_Validation1_revised.xlsx', data_only=True)
ws = wb['Raw_Metrics']

for row in range(3, ws.max_row + 1):
    c_article = ws.cell(row=row, column=2)
    if not isinstance(c_article, MergedCell):
        if not isinstance(ws.cell(row=row, column=4), MergedCell):
            ws.cell(row=row, column=4).value = None # AI Validation Status
        if not isinstance(ws.cell(row=row, column=5), MergedCell):
            ws.cell(row=row, column=5).value = None # Master Code inclusion (we leave this blank for AI to fill)
        if not isinstance(ws.cell(row=row, column=6), MergedCell):
            ws.cell(row=row, column=6).value = None # Exclusion reason
        if not isinstance(ws.cell(row=row, column=16), MergedCell):
            ws.cell(row=row, column=16).value = None # Notes

wb.save('G:\\My Drive\\UCL\\BSMA\\BSMA ANTIGRAVITY\\BSMA_AI_Run_Validation2.xlsx')
print('Created BSMA_AI_Run_Validation2.xlsx')
