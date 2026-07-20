import os
import json
import glob
import openpyxl
import subprocess
import shutil

EXCEL_PATH = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_AI_Run_Validation3.xlsx"
OUTPUTS_DIR = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\scratch\outputs"

def inject_swarm_results():
    brain_dir = r"C:\Users\yunky\.gemini\antigravity\brain"
    target_dir = OUTPUTS_DIR
    
    # 1. Gather scattered JSONs
    if not os.path.exists(target_dir):
        os.makedirs(target_dir)
        
    scattered = glob.glob(os.path.join(brain_dir, "*", "scratch", "outputs", "BSMA*.json"))
    scattered.extend(glob.glob(r"C:\Users\yunky\.gemini\antigravity\scratch\outputs\BSMA*.json"))
    scattered.extend(glob.glob(r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\scratch\outputs\BSMA*.json"))
    for f in scattered:
        try:
            shutil.move(f, os.path.join(target_dir, os.path.basename(f)))
        except Exception as e:
            print(f"Failed to move {f}: {e}")

    # 2. Proceed with injection
    json_files = glob.glob(os.path.join(target_dir, "BSMA*.json"))
    
    if not json_files:
        print(f"Error: No JSON output files found in {target_dir}")
        return
        
    print(f"Found {len(json_files)} results to inject.")
    
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    
    injected_ids = []
    
    for jpath in json_files:
        art_id = os.path.basename(jpath).replace(".json", "")
        
        try:
            try:
                with open(jpath, 'r', encoding='utf-8-sig') as f:
                    verdict = json.load(f)
            except UnicodeDecodeError:
                with open(jpath, 'r', encoding='utf-16') as f:
                    verdict = json.load(f)
            except json.JSONDecodeError:
                # Some files might be totally empty or corrupted.
                print(f"Skipping {jpath} due to JSON Decode Error")
                continue
                
            # Find row
            target_row = None
            for row in range(4, ws.max_row + 1):
                if str(ws.cell(row, 2).value).strip() == art_id:
                    target_row = row
                    break
                    
            if target_row:
                status_val = verdict.get('status')
                if status_val is None:
                    status_val = verdict.get('Verdict', 'ERROR')
                ws.cell(row=target_row, column=5).value = status_val
                
                reason_code_val = verdict.get('reason_code')
                if reason_code_val is None:
                    reason_code_val = verdict.get('Exclusion_Code', '')
                ws.cell(row=target_row, column=6).value = reason_code_val
                
                reason_summary_val = verdict.get('reason_summary')
                if reason_summary_val is None:
                    reason_summary_val = verdict.get('Reason_Summary')
                if reason_summary_val is None:
                    reason_summary_val = verdict.get('Reasoning', '')
                    
                verbatim_val = verdict.get('verbatim')
                if verbatim_val is None:
                    verbatim_val = verdict.get('Verbatim_Evidence', '')
                    
                note_str = f"{reason_summary_val}. Verbatim Evidence: \"{verbatim_val}\""
                ws.cell(row=target_row, column=16).value = note_str
                
                injected_ids.append(art_id)
                # Remove the processed file
                os.remove(jpath)
            else:
                print(f"WARNING: Could not find {art_id} in Excel.")
        except Exception as e:
            print(f"Error processing {art_id}: {e}")
            
    wb.save(EXCEL_PATH)
    print(f"Successfully injected {len(injected_ids)} papers into Excel.")
    
    cwd = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY"
    
    if injected_ids:
        print("\nGenerating Intermediate Batch Report...")
        subprocess.run(["python", ".agents/skills/include_exclude_pipeline/scripts/generate_batch_report.py", "--ids"] + injected_ids, cwd=cwd)
        
    print("\nRunning Post-Highlighting & Backup...")
    subprocess.run(["python", ".agents/skills/include_exclude_pipeline/scripts/post_highlight.py"], cwd=cwd)

if __name__ == "__main__":
    inject_swarm_results()
