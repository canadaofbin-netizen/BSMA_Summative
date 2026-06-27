import openpyxl

excel_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_Actual Coding Sheet_v3.xlsx"
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

def clean_value(val):
    if val is None: return None
    val = str(val).strip()
    # Remove markdown escaped characters
    val = val.replace('\\', '').replace('*', '')
    if val == '999': return 999
    try: return float(val)
    except: return val

def split_compound(val, index):
    """Splits 'BS_val / NonBS_val' and returns the requested index (0 or 1)"""
    if val is None: return 999
    val = str(val).strip()
    if ' / ' in val:
        parts = val.split(' / ')
        if len(parts) > index:
            return clean_value(parts[index])
    return clean_value(val)

# Fix newly inserted rows (Rows 31 to 114)
for r in range(31, 115):
    eff_id = ws.cell(row=r, column=4).value
    if eff_id is None: continue

    # 1. Fix Include column text format
    inc = ws.cell(row=r, column=5).value
    if inc == 1: ws.cell(row=r, column=5, value="1 = Include")
    elif inc == 0: ws.cell(row=r, column=5, value="0 = Exclude")

    # 2. Fix the Columns!
    # In the previous flawed script:
    # Correlate Mean was written to Col 40 (which is actually Notes)
    # Correlate SD to Col 41 (BS Name)
    # Correlate Alpha to Col 42 (BS Mean)
    # r value to Col 45 (Non-BS Name)
    
    # Let's extract the raw flawed strings from these wrong columns
    raw_mean = ws.cell(row=r, column=40).value
    raw_sd = ws.cell(row=r, column=41).value
    raw_alpha = ws.cell(row=r, column=42).value
    raw_r = ws.cell(row=r, column=45).value
    
    # We also need to get the actual names which were stored correctly or incorrectly
    # Actually, it's safer to just re-parse from the Shadow Reports, but we can fix it in-place for Liu 2024.
    
    # For Liu 2024 (BSMA0006):
    if str(eff_id).startswith("BSMA0006"):
        # The previous script did this:
        # ws.cell(row=r_idx, column=36, value=corr_var) # Col 36 is Max Score! (Wrong)
        # It's heavily shifted. It's much better to clear the wrong cells and map them correctly.
        
        corr_var = ws.cell(row=r, column=36).value
        # Clear the wrong cells
        for c in [36, 39, 40, 41, 42, 45]:
            ws.cell(row=r, column=c, value=None)
            
        # Correct Mapping for Liu 2024:
        ws.cell(row=r, column=41, value="Expatriates' boundary-spanning") # BS Name
        ws.cell(row=r, column=42, value=split_compound(raw_mean, 0)) # BS Mean
        ws.cell(row=r, column=43, value=split_compound(raw_sd, 0)) # BS SD
        ws.cell(row=r, column=44, value=split_compound(raw_alpha, 0)) # BS Alpha
        
        ws.cell(row=r, column=45, value=corr_var) # Non-BS Name
        ws.cell(row=r, column=46, value=split_compound(raw_mean, 1)) # Non-BS Mean
        ws.cell(row=r, column=47, value=split_compound(raw_sd, 1)) # Non-BS SD
        ws.cell(row=r, column=48, value=split_compound(raw_alpha, 1)) # Non-BS Alpha
        
        ws.cell(row=r, column=49, value=clean_value(raw_r)) # Correlation r
        
    # For Marrone 2007 (BSMA0001):
    elif str(eff_id).startswith("BSMA0001"):
        # Previous script wrote:
        # ws.cell(row=r_idx, column=36, value=corr_var)
        # ws.cell(row=r_idx, column=42, value=non_bs_alpha)
        # ws.cell(row=r_idx, column=44, value=n_val)
        # ws.cell(row=r_idx, column=45, value=r_val)
        
        corr_var = ws.cell(row=r, column=36).value
        non_bs_alpha = ws.cell(row=r, column=42).value
        raw_r = ws.cell(row=r, column=45).value
        
        for c in [36, 42, 44, 45]:
            ws.cell(row=r, column=c, value=None)
            
        ws.cell(row=r, column=41, value="Boundary-spanning behavior")
        ws.cell(row=r, column=44, value=0.86) # From text
        ws.cell(row=r, column=45, value=corr_var)
        ws.cell(row=r, column=48, value=clean_value(non_bs_alpha))
        ws.cell(row=r, column=49, value=clean_value(raw_r))

wb.save(excel_path)
print("Data cleansing and column realignment completed successfully.")
