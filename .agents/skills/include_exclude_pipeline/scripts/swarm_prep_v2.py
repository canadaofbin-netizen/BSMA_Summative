import os
import json
import openpyxl
import argparse

EXCEL_PATH = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_AI_Run_Validation2.xlsx"
PDF_DIR = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\01_Academic_Papers"
PAYLOAD_OUT = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\scratch\subagents_payload_v2.json"

def find_pdf_for_id(art_id):
    try:
        num = int(art_id.replace("BSMA", ""))
        prefix = f"[{num}]"
        for file in os.listdir(PDF_DIR):
            if file.startswith(prefix) and file.lower().endswith(".pdf"):
                return os.path.join(PDF_DIR, file)
    except Exception:
        pass
    return None

def prep_swarm(batch_size=40):
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    
    pending = []
    for row in range(4, ws.max_row + 1):
        art_id = str(ws.cell(row, 2).value).strip()
        status = ws.cell(row, 5).value
        
        if not status or str(status).strip() == "" or str(status) == "None":
            if art_id and art_id != "None":
                pdf_path = find_pdf_for_id(art_id)
                if pdf_path:
                    pdf_path_clean = pdf_path.replace("\\", "/")
                    pending.append((art_id, pdf_path_clean))
                
                if len(pending) >= batch_size:
                    break
                    
    if not pending:
        print("No more pending papers found!")
        return
        
    print(f"Preparing Swarm for {len(pending)} papers: {pending[0][0]} to {pending[-1][0]}")
    
    subagents = []
    for art_id, pdf_path in pending:
        out_path = os.path.abspath(f"scratch/outputs_v2/{art_id}.json").replace("\\", "/")
        prompt = f"Review `{art_id}`. The PDF is at `{pdf_path}`. Output to `{out_path}`."

        subagents.append({
            "TypeName": "bsma_reviewer_v3",
            "Role": f"Validation2 for {art_id}",
            "Prompt": prompt
        })
        
    os.makedirs(os.path.dirname(PAYLOAD_OUT), exist_ok=True)
    with open(PAYLOAD_OUT, 'w', encoding='utf-8') as f:
        json.dump(subagents, f, indent=4)
        
    print(f"Payload for {len(pending)} subagents saved to {PAYLOAD_OUT}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--batch", type=int, default=40, help="Number of subagents to prep")
    args = parser.parse_args()
    prep_swarm(args.batch)
