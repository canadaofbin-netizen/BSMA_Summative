import openpyxl
import argparse

def grade_test(master_path, test_path):
    wb_master = openpyxl.load_workbook(master_path, data_only=True)
    ws_master = wb_master.active

    wb_test = openpyxl.load_workbook(test_path, data_only=True)
    ws_test = wb_test.active

    master_dict = {}
    for row in range(4, ws_master.max_row + 1):
        art_id = str(ws_master.cell(row, 2).value).strip()
        if art_id and art_id != "None":
            status = str(ws_master.cell(row, 5).value).strip()
            reason = str(ws_master.cell(row, 6).value).strip()
            master_dict[art_id] = {"status": status, "reason": reason}

    correct = 0
    total = 0
    fp = 0
    fn = 0
    mismatches = []
    missing_in_master = []

    for row in range(4, ws_test.max_row + 1):
        art_id = str(ws_test.cell(row, 2).value).strip()
        if art_id and art_id.startswith("BSMA") and art_id != "None":
            test_status = str(ws_test.cell(row, 5).value).strip()
            
            # Skip papers that AI hasn't processed (no status)
            if test_status == "None" or not test_status:
                continue
                
            truth = master_dict.get(art_id)
            if not truth:
                missing_in_master.append(art_id)
                continue
                
            total += 1
            true_status = truth["status"]
            
            test_status_val = "1" if "1" in test_status else ("0" if "0" in test_status else test_status)
            true_status_val = "1" if "1" in true_status else ("0" if "0" in true_status else true_status)
            
            is_correct = True
            
            if test_status_val == true_status_val:
                correct += 1
            else:
                is_correct = False
                if test_status_val == "1" and true_status_val == "0":
                    fp += 1
                elif test_status_val == "0" and true_status_val == "1":
                    fn += 1
                    
            if not is_correct:
                mismatches.append(f"{art_id}: AI='{test_status}' (Truth='{true_status}')")

    print("\n" + "="*40)
    print("FINAL BACKTESTING GRADE REPORT")
    print("="*40)
    print(f"Total Papers Evaluated : {total}")
    if total > 0:
        print(f"Accuracy               : {(correct/total)*100:.2f}% ({correct}/{total})")
    print(f"False Positives (FP)   : {fp}")
    print(f"False Negatives (FN)   : {fn}")
    if mismatches:
        print("\n[Mismatches Detected]")
        for m in mismatches:
            print(f" - {m}")
    if missing_in_master:
        print("\n[Papers Missing in Truth Set]")
        print(missing_in_master)
    print("="*40)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate Match Rate Report between two Excel sheets.")
    parser.add_argument("--master", type=str, required=True, help="Path to the master/truth Excel file")
    parser.add_argument("--test", type=str, required=True, help="Path to the test/generated Excel file")
    args = parser.parse_args()
    
    grade_test(args.master, args.test)
