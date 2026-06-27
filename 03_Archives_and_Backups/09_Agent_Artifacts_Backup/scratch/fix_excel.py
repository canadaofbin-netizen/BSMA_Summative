import openpyxl

excel_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_Actual Coding Sheet_v3.xlsx"
out_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_Actual Coding Sheet_v4.xlsx"
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

max_row = ws.max_row

# 1. Swap Col 1 and Col 2 for newly inserted rows (from row 31 down)
for r in range(31, max_row + 1):
    val1 = ws.cell(row=r, column=1).value
    val2 = ws.cell(row=r, column=2).value
    
    # Only swap if they look like the flipped ones 
    # (val1 = BSMA..., val2 = KY)
    if str(val1).startswith("BSMA") and str(val2) == "KY":
        ws.cell(row=r, column=1, value=val2)
        ws.cell(row=r, column=2, value=val1)

# 2. Insert blank rows between different papers.
# We iterate backwards to avoid shifting row indices during insertion.
prev_article_id = None
# Re-evaluate max_row just in case
for r in range(max_row, 30, -1):
    article_id = ws.cell(row=r, column=2).value
    if article_id is None:
        continue
    
    if prev_article_id is not None and article_id != prev_article_id:
        # The paper changed here, so we insert a blank row BELOW the current row r
        # which is at index r + 1
        ws.insert_rows(r + 1)
        
    prev_article_id = article_id

wb.save(out_path)
print("Fixed Coder Initials/Article ID swap and inserted blank lines between papers.")
