import openpyxl

excel_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_AI_Run_V2.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb.active

with open("scratch/excel_results.txt", "w", encoding="utf-8") as f:
    for row in range(4, ws.max_row+1):
        art_id = ws.cell(row, 2).value
        if art_id in ['BSMA0126', 'BSMA0325', 'BSMA0297', 'BSMA0471', 'BSMA0135', 'BSMA0509', 'BSMA0399', 'BSMA0041', 'BSMA0099']:
            f.write(f"{art_id}: {ws.cell(row, 5).value} | {ws.cell(row, 16).value}\n")
