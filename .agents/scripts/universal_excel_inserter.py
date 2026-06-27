import openpyxl
import os
import argparse
import json

def insert_data(excel_path, data):
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        return

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Text mapping dictionaries
    pub_type_map = {1: "1 = Journal article", 2: "2 = Dissertation/thesis", 3: "3 = Conference paper", 4: "4 = Working paper/unpublished", 5: "5 = Book/book chapter"}
    study_design_map = {1: "1 = Cross-sectional", 2: "2 = Longitudinal/time-lagged", 3: "3 = Experimental", 4: "4 = Qualitative"}
    intl_context_map = {1: "1 = No (domestic only)", 2: "2 = Yes (expatriate, multinational, etc.)"}
    report_type_map = {1: "1 = Self", 2: "2 = Supervisor", 3: "3 = Others", 4: "4 = Objective"}
    inclusion_map = {0: "0 = Exclude", 1: "1 = Include"}

    article_id = data.get("article_id", "BSMA9999")
    inclusion_status = data.get("inclusion_status", 1)
    exclusion_reason = data.get("exclusion_reason", None)

    # Find the target row
    target_row_idx = None
    for r in range(4, ws.max_row + 1):
        if str(ws.cell(row=r, column=2).value).strip() == article_id:
            target_row_idx = r
            break
            
    if target_row_idx is None:
        print(f"Warning: {article_id} not found in existing rows. Appending to bottom.")
        target_row_idx = ws.max_row + 1

    # Extract metadata from the existing row (if it exists) to copy into new rows
    metadata = [ws.cell(row=target_row_idx, column=c).value for c in range(1, 11)] # Cols A to J

    # PATH A: EXCLUSION LOGIC
    if inclusion_status == 0:
        ws.cell(row=target_row_idx, column=5).value = inclusion_map.get(0, "0 = Exclude")
        ws.cell(row=target_row_idx, column=6).value = exclusion_reason
        wb.save(excel_path)
        print(f"Successfully excluded {article_id}. Reason logged in Excel.")
        return

    # PATH B: INCLUSION LOGIC (M:N schema)
    samples = data.get("samples", [])
    if not samples:
        print(f"Error: No samples found in JSON for included paper {article_id}.")
        return

    rows_to_insert = []
    total_rows = 0

    for s_idx, sample in enumerate(samples):
        sample_number = sample.get("sample_number", s_idx + 1)
        sample_id = f"{article_id}.{sample_number}"
        
        sample_size = sample.get("sample_size", None)
        pub_type = pub_type_map.get(sample.get("publication_type"), sample.get("publication_type"))
        study_design = study_design_map.get(sample.get("study_design"), sample.get("study_design"))
        intl_context = intl_context_map.get(sample.get("international_context"), sample.get("international_context"))
        occ_type = sample.get("occupation_type", None)

        bs_measures = sample.get("bs_measures", [])
        correlations = sample.get("correlations", [])

        if not bs_measures or not correlations:
            continue

        for bs_measure in bs_measures:
            bs_name = bs_measure.get("name", "Unknown BS")
            paired_corrs = [c for c in correlations if c.get("bs_name", bs_name) == bs_name]
            if not paired_corrs:
                paired_corrs = correlations # Fallback

            for c_idx, corr in enumerate(paired_corrs):
                row = [None] * 50
                
                # Copy metadata
                for i in range(10):
                    if i < len(metadata):
                        row[i] = metadata[i]
                
                # Identifiers
                row[0] = "KY"
                row[1] = article_id
                row[2] = sample_id
                row[3] = f"{sample_id}.{total_rows + 1}"
                row[4] = inclusion_map.get(1, "1 = Include")
                
                # Categoricals
                row[11] = pub_type
                row[16] = study_design
                row[20] = intl_context
                row[21] = sample_size
                row[25] = occ_type

                # BS Descriptors (Cols 26-32)
                row[26] = bs_measure.get("items")
                row[27] = bs_measure.get("min")
                row[28] = bs_measure.get("max")
                row[29] = report_type_map.get(bs_measure.get("report_type"), bs_measure.get("report_type"))
                row[31] = bs_measure.get("specific_measure")
                row[32] = bs_measure.get("notes")

                # Non-BS Descriptors (Cols 33-39)
                row[33] = corr.get("items")
                row[34] = corr.get("min")
                row[35] = corr.get("max")
                row[36] = report_type_map.get(corr.get("report_type"), corr.get("report_type"))
                row[38] = corr.get("specific_measure")
                row[39] = corr.get("notes")

                # Effect Sizes: BS (Cols 40-43)
                row[40] = bs_name
                row[41] = bs_measure.get("mean")
                row[42] = bs_measure.get("sd")
                row[43] = bs_measure.get("alpha")

                # Effect Sizes: Non-BS (Cols 44-47)
                row[44] = corr.get("non_bs_name")
                row[45] = corr.get("mean")
                row[46] = corr.get("sd")
                row[47] = corr.get("alpha")

                # Correlation
                row[48] = corr.get("r")

                rows_to_insert.append(row)
                total_rows += 1

    # Insert rows into Excel
    if rows_to_insert:
        # First row overwrites the target row
        for col_idx, val in enumerate(rows_to_insert[0], start=1):
            ws.cell(row=target_row_idx, column=col_idx).value = val
            
        # Subsequent rows are inserted below
        if len(rows_to_insert) > 1:
            ws.insert_rows(target_row_idx + 1, len(rows_to_insert) - 1)
            for r_offset, row_data in enumerate(rows_to_insert[1:]):
                for col_idx, val in enumerate(row_data, start=1):
                    ws.cell(row=target_row_idx + 1 + r_offset, column=col_idx).value = val

    wb.save(excel_path)
    print(f"Successfully injected {total_rows} rows into {excel_path} for {article_id}.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Universal Excel Inserter for BSMA")
    parser.add_argument("--excel", type=str, required=True)
    parser.add_argument("--data", type=str, required=True)
    args = parser.parse_args()
    
    try:
        payload = json.loads(args.data)
        insert_data(args.excel, payload)
    except Exception as e:
        print(f"Failed: {e}")
