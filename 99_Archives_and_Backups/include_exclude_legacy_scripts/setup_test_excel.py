import shutil
import openpyxl
import os

master_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_AI_Run_Validation3.xlsx"
test_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\Test_Coding_Sheet.xlsx"

print("Copying Master Sheet to Test Sheet...")
shutil.copy(master_path, test_path)

print("Blanking columns 5 (Status), 6 (Reason Code), 16 (Notes)...")
wb = openpyxl.load_workbook(test_path)
ws = wb.active

for row in range(4, ws.max_row + 1):
    ws.cell(row=row, column=5).value = None
    ws.cell(row=row, column=6).value = None
    ws.cell(row=row, column=16).value = None

wb.save(test_path)
print("Setup complete. Test_Coding_Sheet.xlsx is ready.")
