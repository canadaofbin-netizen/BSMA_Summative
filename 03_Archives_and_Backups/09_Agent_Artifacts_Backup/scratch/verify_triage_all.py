import os
import re
import glob
import openpyxl
import pandas as pd
import fitz
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = 'BSMA_Master_Coding_Sheet.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb.active

df_bq = pd.read_csv('03_Archives_and_Backups/batch_queue.csv')
id_to_fn = dict(zip(df_bq['Article_ID'], df_bq['Filename']))

pdfs = {os.path.basename(p): p for p in glob.glob('01_Academic_Papers/**/*.pdf', recursive=True)}

false_exclusions = []
false_inclusions = []

print("Starting PDF verification across all rows...")

for r in range(2, sheet.max_row + 1):
    art_id = sheet.cell(r, 2).value
    if not art_id:
        continue
    status = sheet.cell(r, 5).value
    title = str(sheet.cell(r, 8).value or '')
    authors = str(sheet.cell(r, 10).value or '')
    year = str(sheet.cell(r, 11).value or '')
    exc_code = str(sheet.cell(r, 6).value or '')
    notes = str(sheet.cell(r, 16).value or '')

    fn = id_to_fn.get(art_id)
    pdf_path = pdfs.get(fn)
    if not pdf_path or not os.path.exists(pdf_path):
        continue

    try:
        doc = fitz.open(pdf_path)
        num_pages = len(doc)
        intro_text = ""
        for i in range(min(3, num_pages)):
            intro_text += doc[i].get_page_text()
        intro_lower = intro_text.lower()
        
        has_corr = False
        has_sample = False
        for i in range(num_pages):
            txt = doc[i].get_page_text()
            txt_l = txt.lower()
            if "correlation" in txt_l or " r =" in txt_l or " r=" in txt_l or "pearson" in txt_l:
                has_corr = True
            if "sample" in txt_l or " n =" in txt_l or " n=" in txt_l or "questionnaire" in txt_l or "respondent" in txt_l:
                has_sample = True
            if has_corr and has_sample:
                break
        doc.close()
    except Exception as e:
        continue

    t_lower = title.lower()
    
    # 1. Check False Exclusions (Currently Excluded = 0)
    if status == '0 = Exclude':
        emp_kw = ['employee', 'salesperson', 'salespeople', 'sales rep', 'nurse', 'teacher', 'boundary spanner', 'expatriate', 'leader', 'manager', 'subordinate', 'staff', 'job satisfaction', 'turnover', 'burnout']
        is_indiv_emp = any(k in t_lower or k in intro_lower[:1500] for k in emp_kw)
        
        non_indiv_kw = ['firm performance', 'team performance', 'sample of firms', 'sample of teams', 'organizational-level', 'firm-level', 'team-level', 'supply chain network', 'inter-organizational relationship']
        is_non_indiv = any(k in t_lower or k in intro_lower[:1500] for k in non_indiv_kw)
        
        if is_indiv_emp and has_corr and has_sample and not is_non_indiv:
            false_exclusions.append((art_id, exc_code, year, authors[:20], title[:60], notes[:40]))

    # 2. Check False Inclusions (Currently Included = 1)
    elif status == '1 = Include':
        non_indiv_kw = ['firm performance', 'team performance', 'sample of firms', 'sample of teams', 'organizational-level', 'firm-level', 'team-level', 'supply chain network', 'inter-organizational relationship', 'business model design', 'pollution control department', 'alliance performance']
        is_non_indiv = any(k in t_lower for k in non_indiv_kw)
        
        non_emp_kw = ['literature review', 'systematic review', 'conceptual framework', 'we propose a theoretical model', 'qualitative case study', 'ethnographic']
        is_non_emp = any(k in t_lower or k in intro_lower[:1000] for k in non_emp_kw) and not has_corr
        
        if is_non_indiv or not has_corr or is_non_emp:
            reason = "Non-Individual Level" if is_non_indiv else ("No Correlation Matrix / Non-Empirical" if not has_corr else "Review/Conceptual")
            false_inclusions.append((art_id, year, authors[:20], title[:60], reason))

print(f"Verification complete! Found {len(false_exclusions)} suspicious False Exclusions and {len(false_inclusions)} suspicious False Inclusions.")

print("\n--- TOP SUSPICIOUS FALSE EXCLUSIONS (Currently Excluded -> Should likely be INCLUDED) ---")
for x in false_exclusions[:25]:
    print(f"{x[0]} | Excl: {x[1]} | {x[2]} | {x[3]} | {x[4]}")

print("\n--- TOP SUSPICIOUS FALSE INCLUSIONS (Currently Included -> Should likely be EXCLUDED) ---")
for x in false_inclusions[:25]:
    print(f"{x[0]} | Reason: {x[4]} | {x[1]} | {x[2]} | {x[3]}")
