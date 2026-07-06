import os
import glob
import openpyxl
import pandas as pd
import fitz
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('BSMA_Master_Coding_Sheet.xlsx', data_only=True)
sheet = wb.active

df_bq = pd.read_csv('03_Archives_and_Backups/batch_queue.csv')
id_to_fn = dict(zip(df_bq['Article_ID'], df_bq['Filename']))

pdfs = {os.path.basename(p): p for p in glob.glob('01_Academic_Papers/**/*.pdf', recursive=True)}

inc_team_firm = []
inc_no_corr = []
exc_with_corr = []

print("Auditing PDFs...")

for r in range(2, sheet.max_row + 1):
    art_id = sheet.cell(r, 2).value
    if not art_id:
        continue
    status = sheet.cell(r, 5).value
    title = str(sheet.cell(r, 8).value or '')
    authors = str(sheet.cell(r, 10).value or '')
    year = str(sheet.cell(r, 11).value or '')
    exc_code = str(sheet.cell(r, 6).value or '')

    fn = id_to_fn.get(art_id)
    pdf_path = pdfs.get(fn)
    if not pdf_path or not os.path.exists(pdf_path):
        continue

    try:
        doc = fitz.open(pdf_path)
        txt = ""
        for i in range(len(doc)):
            txt += doc[i].get_page_text()
        doc.close()
        txt_l = txt.lower()
    except Exception as e:
        continue

    if status == '1 = Include':
        # Check team/firm level
        tf_kw = ['sample of teams', 'sample of firms', 'sample of organizations', 'team boundary spanning', 'firm boundary spanning', 'team-level boundary spanning', 'firm-level boundary spanning', 'teams as boundary spanners', 'multiteam system', 'multi-team system']
        found_tf = [k for k in tf_kw if k in txt_l]
        if found_tf:
            inc_team_firm.append((art_id, year, authors[:15], title[:45], found_tf[:2]))
        
        # Check no correlation
        if 'correlation' not in txt_l and 'pearson' not in txt_l and ' r =' not in txt_l and ' r=' not in txt_l:
            inc_no_corr.append((art_id, year, authors[:15], title[:45]))

    elif status == '0 = Exclude':
        # Check if excluded but has correlation and salespeople/employee boundary spanner
        has_corr = ('correlation' in txt_l or 'pearson' in txt_l) and (' r ' in txt_l or ' r=' in txt_l or 'table ' in txt_l)
        emp_kw = ['salesperson', 'salespeople', 'boundary spanner', 'nurse', 'expatriate', 'customer-contact']
        found_emp = [k for k in emp_kw if k in txt_l]
        tf_kw = ['sample of teams', 'sample of firms', 'team boundary spanning', 'firm boundary spanning']
        is_tf = any(k in txt_l for k in tf_kw)
        
        if has_corr and found_emp and not is_tf:
            exc_with_corr.append((art_id, exc_code, year, authors[:15], title[:45], found_emp[:2]))

print(f"\n==========================================")
print(f"AUDIT RESULTS:")
print(f"1. INCLUDED papers mentioning team/firm boundary spanning or sample of teams/firms: {len(inc_team_firm)}")
print(f"2. INCLUDED papers with NO correlation/pearson words in entire PDF: {len(inc_no_corr)}")
print(f"3. EXCLUDED papers WITH correlation table & salesperson/employee keywords (no team/firm words): {len(exc_with_corr)}")
print(f"==========================================\n")

print("--- TOP FALSE INCLUSIONS (Team/Firm Level words found in Included PDF) ---")
for x in inc_team_firm[:15]:
    print(f"  {x[0]} | {x[1]} | {x[2]} | {x[3]} | Found: {x[4]}")

print("\n--- TOP FALSE INCLUSIONS (No correlation/pearson found in Included PDF) ---")
for x in inc_no_corr[:15]:
    print(f"  {x[0]} | {x[1]} | {x[2]} | {x[3]}")

print("\n--- TOP FALSE EXCLUSIONS (Excluded papers with correlations & employee spanners) ---")
for x in exc_with_corr[:15]:
    print(f"  {x[0]} | Excl Code: {x[1]} | {x[2]} | {x[3]} | {x[4]} | Keywords: {x[5]}")
