import openpyxl
import os

wb = openpyxl.load_workbook('BSMA_Master_Coding_Sheet.xlsx')
ws = wb['Raw_Metrics'] if 'Raw_Metrics' in wb.sheetnames else wb.active
row_idx = None

for row in range(1, ws.max_row + 1):
    if ws.cell(row, 2).value == 'BSMA0280':
        row_idx = row
        break

if row_idx:
    ws.cell(row_idx, 5).value = 0 # Inclusion-Exclusion Judgment
    ws.cell(row_idx, 6).value = 1 # Reason for Exclusion
    ws.cell(row_idx, 16).value = '[Reason summary] No empirical quantitative data; conceptual/descriptive article. Verbatim Evidence: "[Title] Coordinating Cross-Boundary Performance at the FAA" "[About the Author] This article is adapted from Performance Networks: Transforming Governance for the 21st Century, by Lynn Sandra Kahn, 2009"'
    wb.save('BSMA_Master_Coding_Sheet.xlsx')
    print('Excel updated.')
else:
    print('BSMA0280 not found in Excel.')
