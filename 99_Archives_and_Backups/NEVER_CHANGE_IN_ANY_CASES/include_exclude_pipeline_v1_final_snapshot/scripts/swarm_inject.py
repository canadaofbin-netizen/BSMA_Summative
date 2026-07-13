import os
import json
import glob
import openpyxl
import subprocess
import shutil

EXCEL_PATH = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_AI_Run_Validation1.xlsx"
OUTPUTS_DIR = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\scratch\outputs"

def inject_swarm_results():
    brain_dir = r"C:\Users\yunky\.gemini\antigravity\brain"
    target_dir = OUTPUTS_DIR
    
    # 1. Gather scattered JSONs
    if not os.path.exists(target_dir):
        os.makedirs(target_dir)
        
    scattered = glob.glob(os.path.join(brain_dir, "*", "scratch", "outputs", "BSMA*.json"))
    scattered.extend(glob.glob(r"C:\Users\yunky\.gemini\antigravity\scratch\outputs\BSMA*.json"))
    for f in scattered:
        try:
            shutil.move(f, os.path.join(target_dir, os.path.basename(f)))
        except Exception as e:
            print(f"Failed to move {f}: {e}")

    # 2. Proceed with injection
    json_files = glob.glob(os.path.join(target_dir, "BSMA*.json"))
    
    if not json_files:
        print(f"Error: No JSON output files found in {target_dir}")
        return
        
    print(f"Found {len(json_files)} results to inject.")
    
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Raw_Metrics"]
    
    injected_ids = []
    
    for jpath in json_files:
        art_id = os.path.basename(jpath).replace(".json", "")
        
        try:
            try:
                with open(jpath, 'r', encoding='utf-8-sig') as f:
                    verdict = json.load(f)
            except UnicodeDecodeError:
                with open(jpath, 'r', encoding='utf-16') as f:
                    verdict = json.load(f)
            except json.JSONDecodeError:
                # Some files might be totally empty or corrupted.
                print(f"Skipping {jpath} due to JSON Decode Error")
                continue
                
            # Find row
            target_row = None
            for row in range(4, ws.max_row + 1):
                if str(ws.cell(row, 2).value).strip() == art_id:
                    target_row = row
                    break
                    
            if target_row:
                ws.cell(row=target_row, column=5).value = verdict.get('status', 'ERROR')
                ws.cell(row=target_row, column=6).value = verdict.get('reason_code', '')
                note_str = f"{verdict.get('reason_summary', '')}. Verbatim Evidence: \"{verdict.get('verbatim', '')}\""
                ws.cell(row=target_row, column=16).value = note_str
                
                injected_ids.append(art_id)
                # Remove the processed file
                os.remove(jpath)
            else:
                print(f"WARNING: Could not find {art_id} in Excel.")
        except Exception as e:
            print(f"Error processing {art_id}: {e}")
            
    wb.save(EXCEL_PATH)
    print(f"Successfully injected {len(injected_ids)} papers into Excel.")
    
    cwd = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY"
    
    if injected_ids:
        print("\nGenerating Intermediate Batch Report...")
        subprocess.run(["python", ".agents/skills/include_exclude_pipeline/scripts/generate_batch_report.py", "--ids"] + injected_ids, cwd=cwd)
        
    print("\nRunning Post-Highlighting & Backup...")
    subprocess.run(["python", ".agents/skills/include_exclude_pipeline/scripts/post_highlight.py"], cwd=cwd)

if __name__ == "__main__":
    inject_swarm_results()
