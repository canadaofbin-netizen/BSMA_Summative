import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook(r'g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_Actual Coding Sheet_v7.xlsx')
ws = wb.active

# Find BSMA0006 row
liu_row = None
akaho_row = None
marrone_row = None
for r in range(4, ws.max_row + 1):
    val = ws.cell(row=r, column=2).value
    if val:
        if 'BSMA0006' in str(val) and liu_row is None:
            liu_row = r
        elif 'BSMA0007' in str(val) and akaho_row is None:
            akaho_row = r
        elif 'BSMA0008' in str(val) and marrone_row is None:
            marrone_row = r

print("=== BSMA0001 (Row 4) vs BSMA0006 (Liu) COLUMN-BY-COLUMN ===\n")
for col in range(1, 50):
    h = ws.cell(row=3, column=col).value or ws.cell(row=2, column=col).value or ws.cell(row=1, column=col).value or ''
    h = str(h)[:40]
    v1 = ws.cell(row=4, column=col).value
    v2 = ws.cell(row=liu_row, column=col).value if liu_row else ''
    v1_str = str(v1)[:45] if v1 is not None else '(empty)'
    v2_str = str(v2)[:45] if v2 is not None else '(empty)'
    if v1 is not None or v2 is not None:
        match = "OK" if (v1 is None and v2 is None) else ""
        print(f"Col {col:2d} [{h:40s}]")
        print(f"  BSMA0001: {v1_str}")
        print(f"  BSMA0006: {v2_str}")
        print()

print("\n=== BSMA0007 (Akaho - Excluded) ===\n")
if akaho_row:
    for col in range(1, 50):
        val = ws.cell(row=akaho_row, column=col).value
        if val is not None:
            h = ws.cell(row=3, column=col).value or ws.cell(row=2, column=col).value or ws.cell(row=1, column=col).value or ''
            print(f"Col {col:2d} [{h}]: {val}")

print("\n=== BSMA0008 (Marrone - first row) ===\n")
if marrone_row:
    for col in range(1, 50):
        val = ws.cell(row=marrone_row, column=col).value
        if val is not None:
            h = ws.cell(row=3, column=col).value or ws.cell(row=2, column=col).value or ws.cell(row=1, column=col).value or ''
            print(f"Col {col:2d} [{h}]: {val}")
