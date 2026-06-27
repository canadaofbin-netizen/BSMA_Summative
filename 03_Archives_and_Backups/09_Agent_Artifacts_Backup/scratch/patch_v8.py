import openpyxl

wb = openpyxl.load_workbook(r'g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_Actual Coding Sheet_v8.xlsx')
ws = wb.active

for r in range(4, ws.max_row + 1):
    art_id = ws.cell(row=r, column=2).value
    if not art_id: continue
    
    non_bs_items = ws.cell(row=r, column=34).value
    non_bs_var = str(ws.cell(row=r, column=45).value or "").lower()
    
    if non_bs_items is None:
        if 'BSMA0006' in str(art_id) and 'expatriates' in non_bs_var and 'boundary-spanning' in non_bs_var:
            ws.cell(row=r, column=34).value = 26
            ws.cell(row=r, column=35).value = 1
            ws.cell(row=r, column=36).value = 7
            ws.cell(row=r, column=37).value = "3 = Others (e.g. coworker/peer subordinate customer/client)"
            ws.cell(row=r, column=38).value = "host-country coworkers"
            ws.cell(row=r, column=39).value = "Liu et al. (2024)"
            
        elif 'BSMA0008' in str(art_id) and 'boundary-spanning behavior' in non_bs_var:
            ws.cell(row=r, column=34).value = 6
            ws.cell(row=r, column=35).value = 1
            ws.cell(row=r, column=36).value = 5
            ws.cell(row=r, column=37).value = "3 = Others (e.g. coworker/peer subordinate customer/client)"
            ws.cell(row=r, column=38).value = "peer ratings"
            ws.cell(row=r, column=39).value = "Ancona & Caldwell (1992) adapted"

wb.save(r'g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_Actual Coding Sheet_v8.xlsx')
print("Patch applied to v8.xlsx!")
