import openpyxl

wb = openpyxl.load_workbook(r'g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_Actual Coding Sheet_v5.xlsx')
ws = wb.active

empty_cols_tally = {}
headers = {}

for c in range(1, 55):
    h2 = ws.cell(row=2, column=c).value
    h3 = ws.cell(row=3, column=c).value
    headers[c] = f"{h2 or ''} | {h3 or ''}".strip(" |")

# Inspect rows 31 to 87
total_rows = 0
for r in range(31, 88):
    if ws.cell(row=r, column=1).value is None:
        continue # skip blank rows
    total_rows += 1
    for c in range(1, 55):
        val = ws.cell(row=r, column=c).value
        if val is None or str(val).strip() == "":
            empty_cols_tally[c] = empty_cols_tally.get(c, 0) + 1

print(f"Total data rows checked: {total_rows}")
print("Columns with empty values:")
for c, count in sorted(empty_cols_tally.items()):
    if count > 0:
        print(f"Col {c} ({headers.get(c)}): {count} empty rows ({(count/total_rows)*100:.1f}%)")
