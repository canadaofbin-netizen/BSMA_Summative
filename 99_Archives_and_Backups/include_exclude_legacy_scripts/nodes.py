import json
import time
import openpyxl
import os

from graph_state import AgentState

def reviewer_node(state: AgentState) -> AgentState:
    print(f"\n[Reviewer] Analyzing Paper {state['article_id']}...")
    
    # Mock LLM API Call
    # In a real setup: Use langchain or openai SDK here.
    # response = llm.invoke(prompt + state['pdf_text'] + state['auditor_feedback'])
    
    time.sleep(1) # Simulate LLM thinking
    
    # Simulating a JSON output based on BSMA Rules
    verdict = {
        "status": "0 = Exclude",
        "reason_code": "3 = Non-individual level (team/firm/org analysis)",
        "reason_summary": "[Team level]",
        "verbatim": "Data was collected from 82 R&D teams."
    }
    
    state["verdict_json"] = verdict
    state["status_msg"] = "Reviewer completed analysis."
    return state

def auditor_node(state: AgentState) -> AgentState:
    print(f"[Auditor] Inspecting verdict for {state['article_id']}...")
    verdict = state["verdict_json"]
    
    state["iteration_count"] += 1
    
    # Mock LLM API Validation
    time.sleep(1)
    
    # Rule Enforcement check:
    if "Team" in verdict["reason_summary"] and "team" not in verdict["verbatim"].lower():
        state["is_approved"] = False
        state["auditor_feedback"] = "REJECTED: Verbatim does not contain team-level keywords."
        print(f"  -> Auditor Rejected! {state['auditor_feedback']}")
    else:
        state["is_approved"] = True
        state["auditor_feedback"] = "APPROVED"
        print("  -> Auditor Approved!")
        
    return state

def injector_node(state: AgentState) -> AgentState:
    print(f"[Injector] Writing verified data for {state['article_id']} into Excel...")
    verdict = state["verdict_json"]
    excel_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_AI_Run_Validation3.xlsx"
    
    print("\n" + "="*40)
    print(" PENDING INJECTION PLAN ")
    print("="*40)
    print(f"Article ID  : {state['article_id']}")
    print(f"Status      : {verdict.get('status')}")
    print(f"Reason Code : {verdict.get('reason_code')}")
    print(f"Notes       : {verdict.get('reason_summary')}. Verbatim Evidence: \"{verdict.get('verbatim')}\"")
    print("="*40)
    
    auto_inject = os.environ.get("AUTO_INJECT", "0") == "1"
    if not auto_inject:
        confirm = input("Proceed with injection into Excel? (y/n): ")
        if confirm.lower() != 'y':
            print("❌ Injection cancelled by user. Terminating process.")
            state["status_msg"] = "Cancelled by user before injection."
            return state
    
    # Step 1: Fast Data Injection using openpyxl
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    target_row = None
    for row in range(4, ws.max_row + 1):
        if str(ws.cell(row, 2).value).strip() == state["article_id"]:
            target_row = row
            break
            
    if target_row:
        ws.cell(row=target_row, column=5).value = verdict['status']
        ws.cell(row=target_row, column=6).value = verdict['reason_code']
        note_str = f"{verdict['reason_summary']}. Verbatim Evidence: \"{verdict['verbatim']}\""
        ws.cell(row=target_row, column=16).value = note_str
        wb.save(excel_path)

            
    state["status_msg"] = "Successfully injected and formatted."
    print("  -> Injector finished!")
    return state
