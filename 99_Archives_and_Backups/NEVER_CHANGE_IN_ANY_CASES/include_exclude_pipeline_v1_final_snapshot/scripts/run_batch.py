import os
import openpyxl
import time
from datetime import datetime
import subprocess
import sys

# Ensure pipeline directory is in sys.path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from langgraph_main import build_graph

BATCH_SIZE = 40
EXCEL_PATH = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_AI_Run_Validation1.xlsx"
ERROR_REPORT_PATH = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\04_Reports\error_report.md"

def get_pending_papers():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    
    pending = []
    for row in range(4, ws.max_row + 1):
        art_id = str(ws.cell(row, 2).value).strip()
        status = ws.cell(row, 5).value
        
        # If status is empty, None, or whitespace, it's pending
        if not status or str(status).strip() == "" or str(status) == "None":
            if art_id and art_id != "None":
                pending.append(art_id)
                
    return pending

def log_error(article_id, error_msg):
    os.makedirs(os.path.dirname(ERROR_REPORT_PATH), exist_ok=True)
    mode = "a" if os.path.exists(ERROR_REPORT_PATH) else "w"
    with open(ERROR_REPORT_PATH, mode, encoding="utf-8") as f:
        if mode == "w":
            f.write("# 🚨 Pipeline Error Log\n\n")
            f.write("| Article ID | Timestamp | Error Message |\n")
            f.write("|---|---|---|\n")
        
        # Clean up newlines in error message
        clean_err = str(error_msg).replace('\n', ' ').replace('\r', '')
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        f.write(f"| **{article_id}** | {timestamp} | {clean_err} |\n")

def run_orchestrator():
    print("==================================================")
    print("🚀 Starting 701-Paper Non-Stop Orchestrator")
    print("==================================================")
    
    pending_papers = get_pending_papers()
    print(f"Found {len(pending_papers)} pending papers.")
    
    if not pending_papers:
        print("All papers have been processed!")
        return

    # Enable auto-inject for the pipeline nodes
    os.environ["AUTO_INJECT"] = "1"
    
    pipeline = build_graph()
    
    total_processed = 0
    total_errors = 0
    
    for i in range(0, len(pending_papers), BATCH_SIZE):
        batch = pending_papers[i:i + BATCH_SIZE]
        print(f"\n🌊 Starting Wave {i//BATCH_SIZE + 1} (Processing {len(batch)} papers)")
        
        batch_processed_ids = []
        for article_id in batch:
            try:
                print(f"\n--- Processing {article_id} ---")
                
                # In production, PDF text extraction would occur here
                initial_state = {
                    "article_id": article_id,
                    "pdf_text": f"Simulated content for {article_id}",
                    "verdict_json": None,
                    "auditor_feedback": None,
                    "is_approved": False,
                    "iteration_count": 0,
                    "status_msg": "Started"
                }
                
                final_state = pipeline.invoke(initial_state)
                
                if "PENDING HUMAN REVIEW" in final_state["status_msg"]:
                    print(f"⚠️ {article_id} routed to HITL (Failed max retries).")
                    log_error(article_id, "Exceeded 3 retry attempts (HITL routing)")
                    total_errors += 1
                else:
                    batch_processed_ids.append(article_id)
                    total_processed += 1
                    
            except Exception as e:
                print(f"❌ Critical Error processing {article_id}: {str(e)}")
                log_error(article_id, f"Pipeline exception: {str(e)}")
                total_errors += 1
                
        # Generate the intermediate report for this chunk
        if batch_processed_ids:
            print("\n📊 Generating Intermediate Batch Report...")
            try:
                cwd = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY"
                subprocess.run(["python", "05_Pipeline/1_Include_Exclude_Loop/generate_batch_report.py", "--ids"] + batch_processed_ids, check=True, cwd=cwd)
            except Exception as e:
                print(f"⚠️ Failed to generate batch report: {e}")
            
    print("\n==================================================")
    print(f"🏁 All waves completed! Processed: {total_processed}, Errors: {total_errors}")
    print("==================================================")
    
    # Final step: Post Highlighting and Git Backup
    print("\n🖌️ Running Final Post-Highlighting & Backup...")
    try:
        cwd = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY"
        subprocess.run(["python", "05_Pipeline/1_Include_Exclude_Loop/post_highlight.py"], check=True, cwd=cwd)
    except Exception as e:
        print(f"❌ Final Post-Highlighting failed: {e}")

if __name__ == "__main__":
    run_orchestrator()
