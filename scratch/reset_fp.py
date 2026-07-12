import openpyxl

fp_papers = [
    'BSMA0126', 'BSMA0127', 'BSMA0135', 'BSMA0225', 'BSMA0260', 
    'BSMA0298', 'BSMA0303', 'BSMA0325', 'BSMA0339', 'BSMA0358', 
    'BSMA0471', 'BSMA0486', 'BSMA0509', 'BSMA0668', 'BSMA0669',
    # including the 6 that were already pending just in case
    'BSMA0099', 'BSMA0211', 'BSMA0383', 'BSMA0399', 'BSMA0449', 'BSMA0559'
]

excel_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_AI_Run_V2.xlsx"
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

cleared_count = 0
for row in range(4, ws.max_row + 1):
    art_id = str(ws.cell(row=row, column=2).value).strip()
    if art_id in fp_papers:
        ws.cell(row=row, column=5).value = None # Status
        ws.cell(row=row, column=6).value = None # Reason Code
        ws.cell(row=row, column=16).value = None # Notes
        cleared_count += 1

wb.save(excel_path)
print(f"Cleared evaluations for {cleared_count} papers.")
