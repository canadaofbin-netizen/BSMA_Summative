import openpyxl
import shutil

SRC = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_AI_Run_Validation1.xlsx"
DST = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_AI_Run_Validation2.xlsx"

# Copy the structure
shutil.copy2(SRC, DST)

# Clear the data columns (keep Article ID and Article Descriptors)
wb = openpyxl.load_workbook(DST)
ws = wb.active

cleared = 0
for row in range(4, ws.max_row + 1):
    art_id = ws.cell(row, 2).value
    if art_id and str(art_id).strip().startswith("BSMA"):
        # Clear columns 5 (Inclusion/Exclusion), 6 (Reason), 16 (Notes)
        ws.cell(row, 5).value = None
        ws.cell(row, 6).value = None
        ws.cell(row, 16).value = None
        cleared += 1

wb.save(DST)
print(f"Created Validation2 Excel with {cleared} papers (data columns cleared).")
