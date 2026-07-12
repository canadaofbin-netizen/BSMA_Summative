import openpyxl

master_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_Master_Coding_Sheet.xlsx"
ai_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_AI_Run_V2.xlsx"

wb_master = openpyxl.load_workbook(master_path, data_only=True)
ws_master = wb_master.active

wb_ai = openpyxl.load_workbook(ai_path, data_only=True)
ws_ai = wb_ai.active

master_dict = {}
for row in range(4, ws_master.max_row + 1):
    art_id = str(ws_master.cell(row, 2).value).strip()
    if art_id.startswith("BSMA"):
        master_dict[art_id] = {
            "status": str(ws_master.cell(row, 5).value).strip(),
            "reason": str(ws_master.cell(row, 6).value).strip()
        }

ai_dict = {}
for row in range(4, ws_ai.max_row + 1):
    art_id = str(ws_ai.cell(row, 2).value).strip()
    if art_id.startswith("BSMA"):
        ai_dict[art_id] = {
            "status": str(ws_ai.cell(row, 5).value).strip(),
            "reason": str(ws_ai.cell(row, 6).value).strip(),
            "verbatim": str(ws_ai.cell(row, 7).value).strip()
        }

fps = ["BSMA0014", "BSMA0033", "BSMA0041", "BSMA0063", "BSMA0085", "BSMA0099", "BSMA0126", "BSMA0127", "BSMA0135", "BSMA0159", "BSMA0163", "BSMA0180", "BSMA0197", "BSMA0211", "BSMA0213", "BSMA0225", "BSMA0244", "BSMA0252", "BSMA0260", "BSMA0272", "BSMA0279", "BSMA0297", "BSMA0298", "BSMA0303", "BSMA0309", "BSMA0325", "BSMA0339", "BSMA0358", "BSMA0383", "BSMA0399", "BSMA0411", "BSMA0435", "BSMA0441", "BSMA0448", "BSMA0449", "BSMA0463", "BSMA0466", "BSMA0471", "BSMA0478", "BSMA0486", "BSMA0509", "BSMA0543", "BSMA0559", "BSMA0632", "BSMA0663", "BSMA0668", "BSMA0669"]
fns = ["BSMA0043", "BSMA0046", "BSMA0049", "BSMA0054", "BSMA0065", "BSMA0094", "BSMA0234", "BSMA0249", "BSMA0277", "BSMA0304", "BSMA0335", "BSMA0336", "BSMA0349", "BSMA0382", "BSMA0387", "BSMA0445", "BSMA0446", "BSMA0511", "BSMA0591", "BSMA0612", "BSMA0624", "BSMA0654", "BSMA0655", "BSMA0673", "BSMA0680", "BSMA0696"]

with open(r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\scratch\diagnosis.md", "w", encoding="utf-8") as f:
    f.write("=== FALSE POSITIVES (AI=1, Truth=0) ===\n")
    for pid in fps[:7] + fps[-7:]:
        ai_info = ai_dict.get(pid, {})
        truth_info = master_dict.get(pid, {})
        f.write(f"[{pid}] AI Reason: {ai_info.get('reason')}\n")
        f.write(f"    AI Verbatim: {ai_info.get('verbatim')}\n")
        f.write(f"    Truth Reason: {truth_info.get('reason')}\n")
        f.write("-" * 50 + "\n")

    f.write("\n=== FALSE NEGATIVES (AI=0, Truth=1) ===\n")
    for pid in fns[:7] + fns[-7:]:
        ai_info = ai_dict.get(pid, {})
        truth_info = master_dict.get(pid, {})
        f.write(f"[{pid}] AI Reason: {ai_info.get('reason')}\n")
        f.write(f"    AI Verbatim: {ai_info.get('verbatim')}\n")
        f.write(f"    Truth Reason: {truth_info.get('reason')}\n")
        f.write("-" * 50 + "\n")
