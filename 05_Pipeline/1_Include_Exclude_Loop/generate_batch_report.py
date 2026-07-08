import os
import argparse
import openpyxl
from datetime import datetime

def generate_batch_report(excel_path, article_ids, output_dir="04_Reports"):
    os.makedirs(output_dir, exist_ok=True)
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    batch_data = []
    # Row 4 is where data starts in BSMA_Master_Coding_Sheet
    for row in range(4, ws.max_row + 1):
        art_id = str(ws.cell(row, 2).value).strip()
        if art_id in article_ids:
            title_raw = str(ws.cell(row, 3).value).strip()
            title = (title_raw[:40] + "...") if len(title_raw) > 40 else title_raw
            author = str(ws.cell(row, 4).value).strip()
            status = str(ws.cell(row, 5).value).strip()
            reason = str(ws.cell(row, 6).value).strip()
            notes = str(ws.cell(row, 16).value).strip()
            
            batch_data.append({
                "id": art_id,
                "author": author,
                "title": title,
                "status": status,
                "reason": reason,
                "notes": notes
            })
            
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = os.path.join(output_dir, f"Batch_Report_{timestamp}.md")
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(f"# 📊 Batch Processing Report\n\n")
        f.write(f"- **Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"- **Total Papers Processed:** {len(batch_data)}\n\n")
        
        f.write("| Article ID | Author | Status | Reason Code | Notes / Verbatim |\n")
        f.write("|---|---|---|---|---|\n")
        
        for d in batch_data:
            # Clean up newlines so they don't break the markdown table
            clean_notes = d['notes'].replace('\n', '<br>').replace('\r', '')
            clean_reason = d['reason'].replace('\n', ' ').replace('\r', '')
            clean_author = d['author'].replace('\n', ' ').replace('\r', '')
            
            # Use HTML formatting for bold status
            if "Include" in d['status']:
                status_fmt = f"<span style='color:green'>**{d['status']}**</span>"
            else:
                status_fmt = f"<span style='color:red'>**{d['status']}**</span>"
                
            f.write(f"| **{d['id']}** | {clean_author} | {status_fmt} | {clean_reason} | {clean_notes} |\n")
            
    print(f"Batch Report successfully generated at: {report_path}")
    return report_path

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate a Markdown report for a batch of processed papers.")
    parser.add_argument("--ids", nargs='+', required=True, help="List of Article IDs (e.g. BSMA0100 BSMA0101)")
    args = parser.parse_args()
    
    excel_file = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_Master_Coding_Sheet.xlsx"
    generate_batch_report(excel_file, args.ids)
