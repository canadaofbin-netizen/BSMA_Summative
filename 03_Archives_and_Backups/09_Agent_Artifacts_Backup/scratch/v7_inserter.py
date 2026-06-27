"""
v7 Excel Inserter - Corrected Column Mapping
=============================================
Correct 0-indexed mapping (row[index] = Excel Col index+1):
  row[0]  = Col 1:  Coder Initials
  row[1]  = Col 2:  Article ID
  row[2]  = Col 3:  Sample ID
  row[3]  = Col 4:  Effect size ID
  row[4]  = Col 5:  Inclusion/Exclusion Judgment
  row[5]  = Col 6:  Reason for Exclusion
  row[6]  = Col 7:  Abstract
  row[7]  = Col 8:  Title
  row[8]  = Col 9:  Publication Name (Journal)
  row[9]  = Col 10: Authors
  row[10] = Col 11: Year
  row[11] = Col 12: Publication Type
  row[13] = Col 14: Full Citation (APA)
  row[16] = Col 17: Study Design
  row[18] = Col 19: Country
  row[20] = Col 21: International Context
  row[21] = Col 22: Sample size (N)
  row[22] = Col 23: Mean Age
  row[23] = Col 24: % Female
  row[24] = Col 25: Org Tenure
  row[25] = Col 26: Occupation Type
  row[26] = Col 27: Number of Items (BS)
  row[27] = Col 28: Min Score (BS)
  row[28] = Col 29: Max Score (BS)
  row[29] = Col 30: Report Type (BS)
  row[30] = Col 31: Report Type Note (BS)
  row[31] = Col 32: Specific Measure Used (BS)
  row[33] = Col 34: Number of Items (Non-BS)
  row[36] = Col 37: Report Type (Non-BS)
  row[37] = Col 38: Report Type Note (Non-BS)
  row[38] = Col 39: Specific Measure Used (Non-BS)
  row[40] = Col 41: BS Variable Name
  row[41] = Col 42: BS Mean
  row[42] = Col 43: BS SD
  row[43] = Col 44: BS Alpha
  row[44] = Col 45: Non-BS Variable Name
  row[45] = Col 46: Non-BS Mean
  row[46] = Col 47: Non-BS SD
  row[47] = Col 48: Non-BS Alpha
  row[48] = Col 49: Correlation r
"""
import openpyxl
import re
import os

def clean_val(val):
    if val is None: return None
    val = str(val).strip()
    val = val.replace('\\', '').replace('**', '').replace('*', '')
    if val == '999' or val == '999 [FLAGGED]': return 999
    try: return float(val)
    except ValueError: return val

def parse_markdown_table(text):
    tables = re.findall(r'\| Effect Size ID.*?(?=\n\n|\Z)', text, re.DOTALL)
    parsed_tables = []
    for t in tables:
        lines = t.strip().split('\n')
        if len(lines) < 3: continue
        headers = [c.strip().replace('*', '').lower() for c in lines[0].split('|')[1:-1]]
        rows = []
        for line in lines[2:]:
            if not line.startswith('|'): break
            cols = [c.strip() for c in line.split('|')[1:-1]]
            row_dict = dict(zip(headers, cols))
            rows.append(row_dict)
        parsed_tables.extend(rows)
    return parsed_tables

# ===========================
# LIU 2024 DATA
# ===========================
# Journal: Journal of International Business Studies
# Title: Expatriates' boundary-spanning: double-edged effects in multinational enterprises
# Authors: Ting Liu, Tomoki Sekiguchi, Jiayin Qin, Ya Xi Shen
# Year: 2024
# Study Design: 2 = Longitudinal (two waves, 2-week interval)
# Country: China (MNEs based in Laos, Vietnam, Philippines, Malaysia)
# International Context: 2 = Yes (international)
# N = 177 (dyadic pairs from Dataset 3)
# Expatriates: Mage = 33.91, 50.28% female, Msubsidiary_tenure = 2.63
# Host-country coworkers: Mage = 31.84, 51.98% female, Msubsidiary_tenure = 3.37
# BS measure: 26 items (9 functional + 7 linguistic + 10 cultural), 1-7 Likert
# Report type: 3 = Others (host-country coworkers rated expatriates' BS)
# Specific Measure: Liu et al. (2024) - newly developed

LIU_META = {
    "title": "Expatriates' boundary-spanning: double-edged effects in multinational enterprises",
    "journal": "Journal of International Business Studies",
    "authors": "Liu, T., Sekiguchi, T., Qin, J., & Shen, Y. X.",
    "year": 2024,
    "study_design": "2 = Longitudinal",
    "country": "China",
    "intl_context": "2 = Yes (international)",
    "n": 177,
    "mean_age": 33.91,
    "pct_female": 50.28,
    "tenure": 2.63,
    "occupation": "Expatriate employees and host-country coworkers in energy engineering MNEs",
    "bs_items": 26,
    "bs_min": 1,
    "bs_max": 7,
    "bs_report_type": "3 = Others (e.g. coworker/peer subordinate customer/client)",
    "bs_report_note": "host-country coworkers",
    "bs_measure": "Liu et al. (2024) - newly developed",
    "pub_type": "1 = Journal article",
}

# ===========================
# MARRONE 2007 DATA
# ===========================
# Journal: Academy of Management Journal
# Title: A Multilevel Investigation of Antecedents and Consequences of Team Member Boundary-Spanning Behavior
# Authors: Marrone, J. A., Tesluk, P. E., & Carson, J. B.
# Year: 2007
# Study Design: 2 = Longitudinal (measures at multiple time points across semester)
# Country: USA
# International Context: 1 = No (domestic only)
# N = 190 (individual level, 31 teams)
# Demographics: not reported (MBA students)
# BS measure: 6 items, 1-5 scale, peer-rated
# Specific Measure: Ancona & Caldwell (1992) adapted
# Role overload: 3 items, Beehr et al. (1976) adapted

MARRONE_META = {
    "title": "A Multilevel Investigation of Antecedents and Consequences of Team Member Boundary-Spanning Behavior",
    "journal": "Academy of Management Journal",
    "authors": "Marrone, J. A., Tesluk, P. E., & Carson, J. B.",
    "year": 2007,
    "study_design": "2 = Longitudinal",
    "country": "USA",
    "intl_context": "1 = No (domestic only)",
    "n_map": {  # variable-specific Ns from Table 1
        "Boundary-spanning behavior": 190,
        "Boundary-spanning role": 186,
        "Boundary-spanning self-efficacy": 190,
        "Asian ethnicity": 182,
        "GMAT score": 177,
        "Role overload": 186,
    },
    "mean_age": 999,
    "pct_female": 999,
    "tenure": 999,
    "occupation": "MBA students in consulting teams",
    "bs_items": 6,
    "bs_min": 1,
    "bs_max": 5,
    "bs_report_type": "3 = Others (e.g. coworker/peer subordinate customer/client)",
    "bs_report_note": "peer ratings",
    "bs_measure": "Ancona & Caldwell (1992) adapted",
    "pub_type": "1 = Journal article",
}


def append_liu(ws, article_id_num, is_first):
    md_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\03_Shadow_Reports\Liu_2024_Shadow_Report.md"
    with open(md_path, 'r', encoding='utf-8') as f:
        text = f.read()
    
    tables = parse_markdown_table(text)
    article_id_str = f"BSMA{article_id_num:04d}"
    sample_id_str = f"{article_id_str}.1"
    
    if not is_first:
        ws.append([None] * 50)
    
    for i, r_dict in enumerate(tables):
        row = [None] * 50
        effect_id = f"{sample_id_str}.{i+1}"
        
        row[0]  = "KY"                              # Col 1: Coder Initials
        row[1]  = article_id_str                     # Col 2: Article ID
        row[2]  = sample_id_str                      # Col 3: Sample ID
        row[3]  = effect_id                          # Col 4: Effect size ID
        row[4]  = "1 = Include"                      # Col 5: Inclusion
        row[7]  = LIU_META["title"]                  # Col 8: Title
        row[8]  = LIU_META["journal"]                # Col 9: Publication Name
        row[9]  = LIU_META["authors"]                # Col 10: Authors
        row[10] = LIU_META["year"]                   # Col 11: Year
        row[11] = LIU_META["pub_type"]               # Col 12: Publication Type
        row[16] = LIU_META["study_design"]           # Col 17: Study Design
        row[18] = LIU_META["country"]                # Col 19: Country
        row[20] = LIU_META["intl_context"]           # Col 21: International Context
        row[21] = LIU_META["n"]                      # Col 22: Sample size (N)
        row[22] = LIU_META["mean_age"]               # Col 23: Mean Age
        row[23] = LIU_META["pct_female"]             # Col 24: % Female
        row[24] = LIU_META["tenure"]                 # Col 25: Org Tenure
        row[25] = LIU_META["occupation"]             # Col 26: Occupation Type
        row[26] = LIU_META["bs_items"]               # Col 27: # Items (BS)
        row[27] = LIU_META["bs_min"]                 # Col 28: Min Score (BS)
        row[28] = LIU_META["bs_max"]                 # Col 29: Max Score (BS)
        row[29] = LIU_META["bs_report_type"]         # Col 30: Report Type (BS)
        row[30] = LIU_META["bs_report_note"]         # Col 31: Report Type Note
        row[31] = LIU_META["bs_measure"]             # Col 32: Specific Measure (BS)
        
        # Effect Size columns
        row[40] = clean_val(r_dict.get("variable 1"))   # Col 41: BS Variable Name
        row[41] = clean_val(r_dict.get("mean 1"))       # Col 42: BS Mean
        row[42] = clean_val(r_dict.get("sd 1"))         # Col 43: BS SD
        row[43] = clean_val(r_dict.get("alpha 1"))      # Col 44: BS Alpha
        row[44] = clean_val(r_dict.get("variable 2"))   # Col 45: Non-BS Variable Name
        row[45] = clean_val(r_dict.get("mean 2"))       # Col 46: Non-BS Mean
        row[46] = clean_val(r_dict.get("sd 2"))         # Col 47: Non-BS SD
        row[47] = clean_val(r_dict.get("alpha 2"))      # Col 48: Non-BS Alpha
        row[48] = clean_val(r_dict.get("r"))            # Col 49: Correlation r
        
        ws.append(row)
    
    print(f"  -> Inserted {len(tables)} rows for Liu 2024")


def append_akaho(ws, article_id_num, is_first):
    article_id_str = f"BSMA{article_id_num:04d}"
    sample_id_str = f"{article_id_str}.1"
    effect_id = f"{sample_id_str}.1"
    
    if not is_first:
        ws.append([None] * 50)
    
    row = [None] * 50
    row[0]  = "KY"                                   # Col 1: Coder Initials
    row[1]  = article_id_str                          # Col 2: Article ID
    row[2]  = sample_id_str                           # Col 3: Sample ID
    row[3]  = effect_id                               # Col 4: Effect size ID (FIXED!)
    row[4]  = "0 = Exclude"                           # Col 5: Exclusion
    row[5]  = "4 = Non-primary study / Theoretical paper" # Col 6: Reason
    
    ws.append(row)
    print(f"  -> Inserted 1 row for Akaho 2024 (Excluded)")


def append_marrone(ws, article_id_num, is_first):
    md_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\03_Shadow_Reports\Marrone_2007_Shadow_Report.md"
    with open(md_path, 'r', encoding='utf-8') as f:
        text = f.read()
    
    tables = parse_markdown_table(text)
    article_id_str = f"BSMA{article_id_num:04d}"
    sample_id_str = f"{article_id_str}.1"
    
    if not is_first:
        ws.append([None] * 50)
    
    for i, r_dict in enumerate(tables):
        row = [None] * 50
        effect_id = f"{sample_id_str}.{i+1}"
        
        n_val = clean_val(r_dict.get("sample size"))
        
        row[0]  = "KY"                                  # Col 1: Coder Initials
        row[1]  = article_id_str                          # Col 2: Article ID
        row[2]  = sample_id_str                           # Col 3: Sample ID
        row[3]  = effect_id                               # Col 4: Effect size ID (FIXED!)
        row[4]  = "1 = Include"                           # Col 5: Inclusion
        row[7]  = MARRONE_META["title"]                   # Col 8: Title
        row[8]  = MARRONE_META["journal"]                 # Col 9: Publication Name
        row[9]  = MARRONE_META["authors"]                 # Col 10: Authors
        row[10] = MARRONE_META["year"]                    # Col 11: Year
        row[11] = MARRONE_META["pub_type"]                # Col 12: Publication Type
        row[16] = MARRONE_META["study_design"]            # Col 17: Study Design
        row[18] = MARRONE_META["country"]                 # Col 19: Country
        row[20] = MARRONE_META["intl_context"]            # Col 21: International Context
        row[21] = n_val                                   # Col 22: Sample size (N)
        row[22] = MARRONE_META["mean_age"]                # Col 23: Mean Age
        row[23] = MARRONE_META["pct_female"]              # Col 24: % Female
        row[24] = MARRONE_META["tenure"]                  # Col 25: Org Tenure
        row[25] = MARRONE_META["occupation"]              # Col 26: Occupation Type
        row[26] = MARRONE_META["bs_items"]                # Col 27: # Items (BS)
        row[27] = MARRONE_META["bs_min"]                  # Col 28: Min Score (BS)
        row[28] = MARRONE_META["bs_max"]                  # Col 29: Max Score (BS)
        row[29] = MARRONE_META["bs_report_type"]          # Col 30: Report Type (BS)
        row[30] = MARRONE_META["bs_report_note"]          # Col 31: Report Type Note
        row[31] = MARRONE_META["bs_measure"]              # Col 32: Specific Measure (BS)
        
        # Effect Size columns
        row[40] = clean_val(r_dict.get("variable 1"))     # Col 41: BS Variable Name
        row[41] = clean_val(r_dict.get("mean 1"))         # Col 42: BS Mean
        row[42] = clean_val(r_dict.get("sd 1"))           # Col 43: BS SD
        row[43] = clean_val(r_dict.get("alpha 1"))        # Col 44: BS Alpha
        row[44] = clean_val(r_dict.get("variable 2"))     # Col 45: Non-BS Variable Name
        row[45] = clean_val(r_dict.get("mean 2"))         # Col 46: Non-BS Mean
        row[46] = clean_val(r_dict.get("sd 2"))           # Col 47: Non-BS SD
        row[47] = clean_val(r_dict.get("alpha 2"))        # Col 48: Non-BS Alpha
        row[48] = clean_val(r_dict.get("r"))              # Col 49: Correlation r
        
        ws.append(row)
    
    print(f"  -> Inserted {len(tables)} rows for Marrone 2007")


if __name__ == "__main__":
    v2_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_Actual Coding Sheet_v2.xlsx"
    v7_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_Actual Coding Sheet_v7.xlsx"
    
    wb = openpyxl.load_workbook(v2_path)
    ws = wb.active
    
    max_id = 0
    for r in range(4, ws.max_row + 1):
        val = ws.cell(row=r, column=2).value
        if val and str(val).startswith('BSMA'):
            m = re.search(r'\d+', str(val))
            if m: max_id = max(max_id, int(m.group()))
    
    print(f"Base Max Article ID: BSMA{max_id:04d}")
    
    # Liu 2024
    max_id += 1
    print(f"\nAppending Liu 2024 as BSMA{max_id:04d}...")
    append_liu(ws, max_id, is_first=True)
    
    # Akaho 2024
    max_id += 1
    print(f"\nAppending Akaho 2024 as BSMA{max_id:04d}...")
    append_akaho(ws, max_id, is_first=False)
    
    # Marrone 2007
    max_id += 1
    print(f"\nAppending Marrone 2007 as BSMA{max_id:04d}...")
    append_marrone(ws, max_id, is_first=False)
    
    wb.save(v7_path)
    print(f"\n=== Generated {v7_path} successfully! ===")
