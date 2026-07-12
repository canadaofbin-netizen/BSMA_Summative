import os
import json
import glob
import openpyxl
import argparse

EXCEL_PATH = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_AI_Run_V2.xlsx"
PDF_DIR = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\01_Academic_Papers"
PAYLOAD_OUT = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\scratch\subagents_payload.json"

def find_pdf_for_id(art_id):
    try:
        num = int(art_id.replace("BSMA", ""))
        prefix = f"[{num}]"
        for file in os.listdir(PDF_DIR):
            if file.startswith(prefix) and file.lower().endswith(".pdf"):
                return os.path.join(PDF_DIR, file)
    except Exception:
        pass
    return None

def prep_swarm(batch_size=20):
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    
    pending = []
    for row in range(4, ws.max_row + 1):
        art_id = str(ws.cell(row, 2).value).strip()
        status = ws.cell(row, 5).value
        
        if not status or str(status).strip() == "" or str(status) == "None":
            if art_id and art_id != "None":
                pdf_path = find_pdf_for_id(art_id)
                if pdf_path:
                    # Convert to forward slashes for prompt robustness
                    pdf_path_clean = pdf_path.replace("\\", "/")
                    pending.append((art_id, pdf_path_clean))
                else:
                    print(f"WARNING: Missing PDF file for {art_id}. Skipping.")
                
                if len(pending) >= batch_size:
                    break
                    
    if not pending:
        print("No more pending papers found!")
        return
        
    print(f"Preparing Swarm for {len(pending)} papers: {pending[0][0]} to {pending[-1][0]}")
    
    subagents = []
    for art_id, pdf_path in pending:
        prompt = f"Review `{art_id}`. The PDF is at `{pdf_path}`. Output to `scratch/outputs/{art_id}.json`."

        subagents.append({
            "TypeName": "bsma_reviewer_v5",
            "Role": f"Reviewer for {art_id}",
            "Prompt": prompt
        })
        
    os.makedirs(os.path.dirname(PAYLOAD_OUT), exist_ok=True)
    with open(PAYLOAD_OUT, 'w', encoding='utf-8') as f:
        json.dump(subagents, f, indent=4)
        
    print(f"Payload for {len(pending)} subagents saved to {PAYLOAD_OUT}")
    print(f"-> Antigravity: You can now read this file and pass it to `invoke_subagent`.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--batch", type=int, default=20, help="Number of subagents to prep")
    args = parser.parse_args()
    prep_swarm(args.batch)
