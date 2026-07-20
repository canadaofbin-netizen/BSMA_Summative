"""
swarm_inject_v3.py — Security Plan v2 적용
==========================================
Layer 2: 오염 검사 (Contamination Detection)
Layer 2: 절단 감지 (Truncation Detection)
Layer 3: 검역 격리 (Quarantine Containment)
Layer 3: 삼중 잠금 주입 (Triple-Lock Injection)
Layer 4: JSON 자동 복원 (JSON Auto-Repair)
Layer 5: 무결성 보고서 (Integrity Report)
"""

import os
import json
import glob
import re
import csv
import shutil
import openpyxl
import subprocess
import argparse

# Removed hardcoded EXCEL_PATH
OUTPUTS_DIR = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\scratch\outputs"
QUARANTINE_DIR = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\scratch\quarantine"
CWD = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY"

# ============================================================
# Layer 2-1: Contamination Detection (오염 패턴 검사)
# ============================================================
CONTAMINATION_PATTERNS = [
    r'</SYSTEM_MESSAGE>',
    r'<SYSTEM_MESSAGE>',
    r'</PLANNER_RESPONSE>',
    r'<PLANNER_RESPONSE>',
    r'\\n',              # escaped newline (not real newline)
    r'\\"',              # escaped quote
    r'<truncated \d+ bytes>',
    r'"step_index"',     # JSON structure key bleed
    r'"tool_calls"',     # JSON structure key bleed
    r'"source":\s*"MODEL"',
]

def check_contamination(text, bsma_id):
    """Pre-injection contamination scan. Returns list of detected patterns."""
    found = []
    for pattern in CONTAMINATION_PATTERNS:
        if re.search(pattern, text):
            found.append(pattern)
    if found:
        print(f"  [CONTAMINATION] {bsma_id}: Detected {len(found)} pattern(s): {found}")
    return found

# ============================================================
# Layer 2-2: Truncation Detection (절단 감지)
# ============================================================
def detect_truncation(raw_text, bsma_id):
    """Detect if the JSON output was truncated by token limits."""
    issues = []

    # CHECK 1: Brace/bracket balance
    open_braces = raw_text.count('{') - raw_text.count('}')
    open_brackets = raw_text.count('[') - raw_text.count(']')
    if open_braces != 0:
        issues.append(f"unbalanced_braces({open_braces:+d})")
    if open_brackets != 0:
        issues.append(f"unbalanced_brackets({open_brackets:+d})")

    # CHECK 2: Odd unescaped quotes
    # Count quotes that are NOT preceded by backslash
    unescaped = len(re.findall(r'(?<!\\)"', raw_text))
    if unescaped % 2 != 0:
        issues.append("odd_quotes")

    # CHECK 3: Missing critical fields
    for field in ['BSMA_ID', 'Verdict', 'Exclusion_Code']:
        if f'"{field}"' not in raw_text and f"'{field}'" not in raw_text:
            issues.append(f"missing_{field}")

    if issues:
        print(f"  [TRUNCATION] {bsma_id}: Detected issues: {issues}")
    return issues

# ============================================================
# Layer 4: JSON Auto-Repair (절단된 JSON 복원)
# ============================================================
def attempt_json_repair(raw_text, bsma_id):
    """Attempt to repair truncated JSON. Returns parsed dict or None."""
    repaired = raw_text.rstrip()

    # Step 1: Close unclosed string
    unescaped = len(re.findall(r'(?<!\\)"', repaired))
    if unescaped % 2 != 0:
        repaired += '"'

    # Step 2: Close unclosed brackets/braces
    open_brackets = repaired.count('[') - repaired.count(']')
    open_braces = repaired.count('{') - repaired.count('}')
    repaired += ']' * max(0, open_brackets)
    repaired += '}' * max(0, open_braces)

    try:
        data = json.loads(repaired)
        # Verify critical fields survived
        if 'Verdict' in data and 'Exclusion_Code' in data:
            # Mark truncated evidence
            if 'Verbatim_Evidence' in data:
                evidence = data['Verbatim_Evidence']
                if isinstance(evidence, str):
                    data['Verbatim_Evidence'] = evidence + ' [TRUNCATED BY TOKEN LIMIT]'
            print(f"  [REPAIR SUCCESS] {bsma_id}: JSON repaired successfully.")
            return data
        else:
            print(f"  [REPAIR FAILED] {bsma_id}: Critical fields missing after repair.")
            return None
    except json.JSONDecodeError as e:
        print(f"  [REPAIR FAILED] {bsma_id}: {e}")
        return None

# ============================================================
# Layer 2-3: Schema Validation (스키마 검증)
# ============================================================
REQUIRED_SCHEMA = {
    "BSMA_ID": str,
    "Verdict": int,
    "Exclusion_Code": (int, type(None)),
    "Reasoning": str,
    "Verbatim_Evidence": str,
}

# Also accept alternate key names from the existing pipeline
KEY_ALIASES = {
    "status": "Verdict",
    "reason_code": "Exclusion_Code",
    "reason_summary": "Reasoning",
    "verbatim": "Verbatim_Evidence",
}

def normalize_keys(data):
    """Normalize alternate key names to canonical schema."""
    normalized = {}
    for k, v in data.items():
        canonical = KEY_ALIASES.get(k, k)
        normalized[canonical] = v
    return normalized

def validate_schema(data, bsma_id):
    """Validate that all required fields exist and have correct types."""
    for key, expected_type in REQUIRED_SCHEMA.items():
        if key not in data:
            print(f"  [SCHEMA] {bsma_id}: Missing field '{key}'")
            return False
        if not isinstance(data[key], expected_type):
            # Allow flexible verdict values
            if key == "Verdict" and isinstance(data[key], str):
                continue  # Will be handled during injection
            print(f"  [SCHEMA] {bsma_id}: Field '{key}' type mismatch")
            return False
    return True

# ============================================================
# Layer 3: Quarantine (검역 격리)
# ============================================================
def move_to_quarantine(json_path, reason, bsma_id):
    """Move contaminated/broken file to quarantine directory."""
    os.makedirs(QUARANTINE_DIR, exist_ok=True)
    dest = os.path.join(QUARANTINE_DIR, os.path.basename(json_path))
    shutil.move(json_path, dest)

    # Log to quarantine CSV
    log_path = os.path.join(QUARANTINE_DIR, "quarantine_log.csv")
    file_exists = os.path.exists(log_path)
    with open(log_path, 'a', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(["Timestamp", "BSMA_ID", "Reason", "File"])
        writer.writerow([datetime.now().isoformat(), bsma_id, reason, os.path.basename(json_path)])

    print(f"  [QUARANTINE] {bsma_id}: Moved to quarantine. Reason: {reason}")

# ============================================================
# Main: Triple-Lock Injection Pipeline
# ============================================================
def inject_swarm_results(excel_path):
    brain_dir = r"C:\Users\yunky\.gemini\antigravity\brain"
    target_dir = OUTPUTS_DIR

    # 0. Gather scattered JSONs from subagent workspaces
    os.makedirs(target_dir, exist_ok=True)
    scattered = glob.glob(os.path.join(brain_dir, "*", "scratch", "outputs", "BSMA*.json"))
    scattered.extend(glob.glob(r"C:\Users\yunky\.gemini\antigravity\scratch\outputs\BSMA*.json"))
    for f in scattered:
        try:
            shutil.move(f, os.path.join(target_dir, os.path.basename(f)))
        except Exception as e:
            print(f"  Failed to move {f}: {e}")

    # 1. Collect all JSON files
    json_files = glob.glob(os.path.join(target_dir, "BSMA*.json"))
    if not json_files:
        print(f"Error: No JSON output files found in {target_dir}")
        return

    print(f"Found {len(json_files)} results to process.")

    # Integrity report counters
    stats = {
        "total": len(json_files),
        "injected_clean": 0,
        "injected_repaired": 0,
        "quarantined_contamination": 0,
        "quarantined_schema": 0,
        "quarantined_unrecoverable": 0,
    }
    quarantined_ids = []
    repaired_ids = []

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    injected_ids = []

    for jpath in json_files:
        art_id = os.path.basename(jpath).replace(".json", "")
        print(f"\n--- Processing {art_id} ---")

        try:
            with open(jpath, 'r', encoding='utf-8-sig') as f:
                raw_text = f.read()
        except UnicodeDecodeError:
            try:
                with open(jpath, 'r', encoding='utf-16') as f:
                    raw_text = f.read()
            except Exception:
                print(f"  [ERROR] {art_id}: Cannot read file.")
                move_to_quarantine(jpath, "unreadable_file", art_id)
                stats["quarantined_unrecoverable"] += 1
                quarantined_ids.append(art_id)
                continue

        # ========== LOCK 1: Truncation Detection ==========
        trunc_issues = detect_truncation(raw_text, art_id)

        if trunc_issues:
            # Attempt JSON repair
            data = attempt_json_repair(raw_text, art_id)
            if data is None:
                move_to_quarantine(jpath, f"truncation_unrecoverable:{trunc_issues}", art_id)
                stats["quarantined_unrecoverable"] += 1
                quarantined_ids.append(art_id)
                continue
            was_repaired = True
        else:
            # Normal JSON parse
            try:
                data = json.loads(raw_text)
                was_repaired = False
            except json.JSONDecodeError as e:
                print(f"  [JSON ERROR] {art_id}: {e}")
                # Last resort: try repair even without detected truncation
                data = attempt_json_repair(raw_text, art_id)
                if data is None:
                    move_to_quarantine(jpath, f"json_parse_failure:{e}", art_id)
                    stats["quarantined_unrecoverable"] += 1
                    quarantined_ids.append(art_id)
                    continue
                was_repaired = True

        # Normalize key names
        data = normalize_keys(data)

        # ========== LOCK 2: Schema Validation ==========
        if not validate_schema(data, art_id):
            move_to_quarantine(jpath, "schema_violation", art_id)
            stats["quarantined_schema"] += 1
            quarantined_ids.append(art_id)
            continue

        # ========== LOCK 3: Contamination Check ==========
        contamination_found = False
        for field in ['Reasoning', 'Verbatim_Evidence']:
            val = data.get(field, '')
            if isinstance(val, str) and check_contamination(val, art_id):
                contamination_found = True
                break

        if contamination_found:
            move_to_quarantine(jpath, "text_contamination", art_id)
            stats["quarantined_contamination"] += 1
            quarantined_ids.append(art_id)
            continue

        # ========== ALL LOCKS PASSED: Inject to Excel ==========
        target_row = None
        for row in range(4, ws.max_row + 1):
            if str(ws.cell(row, 2).value).strip() == art_id:
                target_row = row
                break

        if target_row:
            # Map Verdict to status string
            verdict_val = data.get('Verdict', '')
            if isinstance(verdict_val, int):
                status_str = "1 = Include" if verdict_val == 1 else "0 = Exclude"
            else:
                status_str = str(verdict_val)

            ws.cell(row=target_row, column=5).value = status_str
            ws.cell(row=target_row, column=6).value = data.get('Exclusion_Code', '')
            note_str = f"{data.get('Reasoning', '')}. Verbatim Evidence: \"{data.get('Verbatim_Evidence', '')}\""
            ws.cell(row=target_row, column=16).value = note_str

            injected_ids.append(art_id)
            if was_repaired:
                stats["injected_repaired"] += 1
                repaired_ids.append(art_id)
            else:
                stats["injected_clean"] += 1

            # Remove processed file
            os.remove(jpath)
            print(f"  [INJECTED] {art_id}: {'(repaired)' if was_repaired else '(clean)'}")
        else:
            print(f"  WARNING: Could not find {art_id} in Excel.")

    wb.save(excel_path)

    # ============================================================
    # Layer 5: Integrity Report
    # ============================================================
    total_injected = stats["injected_clean"] + stats["injected_repaired"]
    total_quarantined = stats["quarantined_contamination"] + stats["quarantined_schema"] + stats["quarantined_unrecoverable"]
    contamination_rate = (total_quarantined / stats["total"] * 100) if stats["total"] > 0 else 0

    print("\n" + "=" * 60)
    print("  BATCH INTEGRITY REPORT")
    print("=" * 60)
    print(f"  Total papers processed:         {stats['total']}")
    print(f"  Successfully injected (clean):  {stats['injected_clean']}")
    print(f"  Successfully injected (repaired):{stats['injected_repaired']}  {repaired_ids if repaired_ids else ''}")
    print(f"  Quarantined (contamination):    {stats['quarantined_contamination']}")
    print(f"  Quarantined (schema fail):      {stats['quarantined_schema']}")
    print(f"  Quarantined (unrecoverable):    {stats['quarantined_unrecoverable']}")
    print(f"  ---")
    print(f"  Total Injection Rate:           {total_injected}/{stats['total']} ({total_injected/stats['total']*100:.1f}%)" if stats['total'] > 0 else "  N/A")
    print(f"  Contamination Rate:             {contamination_rate:.1f}%  (threshold: < 10%)")
    status_str = "✅ PASS" if contamination_rate < 10 else "❌ FAIL — PIPELINE HALT RECOMMENDED"
    print(f"  Status:                         {status_str}")
    print("=" * 60)

    if contamination_rate >= 10:
        print("\n  ⚠️  WARNING: Contamination rate exceeds 10% threshold!")
        print("  ⚠️  Review quarantined files before continuing.\n")

    print(f"\nSuccessfully injected {total_injected} papers into Excel.")

    # Generate batch report & post-highlight
    if injected_ids:
        print("\nGenerating Intermediate Batch Report...")
        subprocess.run(["python", ".agents/skills/include_exclude_pipeline/scripts/generate_batch_report.py", "--ids"] + injected_ids, cwd=CWD)

    print("\nRunning Post-Highlighting & Backup...")
    subprocess.run(["python", ".agents/skills/include_exclude_pipeline/scripts/post_highlight.py"], cwd=CWD)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", type=str, required=True, help="Absolute path to the target Excel file")
    args = parser.parse_args()
    inject_swarm_results(args.excel)
