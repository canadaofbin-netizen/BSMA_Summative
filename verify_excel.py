import openpyxl

wb = openpyxl.load_workbook('BSMA_Master_Coding_Sheet.xlsx')
ws = wb.active

# Print the last 12 rows for BSMA0008
print("Verifying inserted rows for BSMA0008:")
for r in range(ws.max_row - 11, ws.max_row + 1):
    row_data = [ws.cell(row=r, column=c).value for c in range(1, 50)]
    # Just print the article id, sample id, bs_name, non_bs_name, and correlation (col 49)
    print(f"Row {r}: Article {row_data[1]}, Sample {row_data[2]}, BS: {row_data[39]}, Non-BS: {row_data[43]}, r: {row_data[48]}")
