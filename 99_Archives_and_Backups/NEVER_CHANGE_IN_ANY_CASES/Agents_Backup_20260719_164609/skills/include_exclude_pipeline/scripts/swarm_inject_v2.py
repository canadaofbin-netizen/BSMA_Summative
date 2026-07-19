import os
import json
import shutil
import openpyxl

EXCEL_PATH = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_AI_Run_Validation2.xlsx"
OUTPUTS_DIR = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\scratch\outputs_v2"
FAILED_DIR = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\scratch\failed_v2"

def inject_swarm_results():
    if not os.path.exists(OUTPUTS_DIR):
        print(f"Directory not found: {OUTPUTS_DIR}")
        return

    os.makedirs(FAILED_DIR, exist_ok=True)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Sheet1']
    
    id_to_row = {}
    for row in range(3, ws.max_row + 1):
        val = ws.cell(row, 2).value
        if val:
            id_to_row[str(val).strip()] = row

    injected = 0
    failed = 0
    not_found = 0

    for filename in os.listdir(OUTPUTS_DIR):
        if filename.endswith(".json"):
            art_id = filename.replace(".json", "")
            filepath = os.path.join(OUTPUTS_DIR, filename)
            
            # Try parsing JSON with multiple encodings
            data = None
            for encoding in ['utf-8-sig', 'utf-8', 'utf-16']:
                try:
                    with open(filepath, 'r', encoding=encoding) as f:
                        content = f.read()
                    data = json.loads(content)
                    break
                except Exception:
                    continue
            
            if data is None:
                # Quarantine failed JSON for manual review
                failed += 1
                dest = os.path.join(FAILED_DIR, filename)
                shutil.move(filepath, dest)
                print(f"QUARANTINED: {art_id} -> {FAILED_DIR} (JSON parse failure)")
                continue
            
            if art_id in id_to_row:
                row = id_to_row[art_id]
                status = data.get("status", "")
                reason = data.get("reason_code", "")
                summary = data.get("reason_summary", "")
                verbatim = data.get("verbatim", "")
                
                status_text = "1 = Include" if str(status) == "1" else "0 = Exclude"
                
                mapping = {
                    '1': '1 = No effect size of interest',
                    '2': '2 = Non-employee samples',
                    '3': '3 = Non-individual level (team/firm/org analysis)',
                    '4': '4 = Non-primary study',
                    '5': '5 = Multiple reasons (specify in Notes)',
                    '6': '6 = Duplicate',
                    '7': '7 = Non-English language',
                    '99': '99 = Other (specify in Notes)'
                }
                
                if reason and str(reason).strip() in mapping:
                    reason = mapping[str(reason).strip()]

                ws.cell(row, 5).value = status_text
                ws.cell(row, 6).value = reason
                ws.cell(row, 16).value = f"{summary}. Verbatim Evidence: {verbatim}"
                injected += 1
                
                os.remove(filepath)
            else:
                not_found += 1
                print(f"WARNING: {art_id} not found in Excel. Keeping JSON.")
                
    wb.save(EXCEL_PATH)
    print(f"\n=== Injection Summary ===")
    print(f"Successfully injected: {injected}")
    print(f"Quarantined (parse failure): {failed}")
    print(f"Not found in Excel: {not_found}")

if __name__ == "__main__":
    inject_swarm_results()

