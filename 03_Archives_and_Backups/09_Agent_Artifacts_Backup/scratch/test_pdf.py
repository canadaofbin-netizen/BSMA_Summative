import os
import glob
import openpyxl
import pandas as pd
import fitz
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('BSMA_Master_Coding_Sheet.xlsx', data_only=True)
sheet = wb.active

df_bq = pd.read_csv('03_Archives_and_Backups/batch_queue.csv')
id_to_fn = dict(zip(df_bq['Article_ID'], df_bq['Filename']))

pdfs = {os.path.basename(p): p for p in glob.glob('01_Academic_Papers/**/*.pdf', recursive=True)}

for r in range(4, 15):
    art_id = sheet.cell(r, 2).value
    status = sheet.cell(r, 5).value
    fn = id_to_fn.get(art_id)
    pdf_path = pdfs.get(fn)
    if not pdf_path:
        print(art_id, "NO PDF FOUND")
        continue
    doc = fitz.open(pdf_path)
    txt = ""
    for i in range(min(3, len(doc))):
        txt += doc[i].get_page_text()
    doc.close()
    txt_l = txt.lower()
    print(f"{art_id} | {status} | len={len(txt)} | corr={'correlation' in txt_l} | team={'team' in txt_l} | firm={'firm' in txt_l}")
