import openpyxl
import re

CONTAMINATION_PATTERNS = [
    r'</SYSTEM_MESSAGE>',
    r'<SYSTEM_MESSAGE>',
    r'</PLANNER_RESPONSE>',
    r'<PLANNER_RESPONSE>',
    r'\\n',              
    r'\\"',              
    r'<truncated \d+ bytes>',
    r'"step_index"',     
    r'"tool_calls"',     
    r'"source":\s*"MODEL"',
]

FILES_TO_CHECK = [
    r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_AI_Run_Validation1.xlsx",
    r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_AI_Run_Validation2.xlsx"
]

def scan_excel(filepath):
    print(f"\n========================================================")
    print(f"Scanning: {filepath}")
    print(f"========================================================")
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"Failed to open {filepath}: {e}")
        return
        
    contamination_count = 0
    truncation_suspicions = 0
    empty_notes_count = 0
    total_processed = 0
    
    for row in range(4, ws.max_row + 1):
        art_id = str(ws.cell(row, 2).value).strip()
        status = ws.cell(row, 5).value
        notes = ws.cell(row, 16).value
        
        # Only check rows that have been processed by AI
        if status and str(status).strip() != "" and str(status) != "None":
            total_processed += 1
            
            if not notes or str(notes).strip() == "" or str(notes) == "None":
                empty_notes_count += 1
                print(f"  [EMPTY] {art_id}: Processed but Notes column is empty.")
                continue
                
            notes_str = str(notes)
            
            # Check for contamination
            found_patterns = []
            for pattern in CONTAMINATION_PATTERNS:
                if re.search(pattern, notes_str):
                    found_patterns.append(pattern)
            
            if found_patterns:
                contamination_count += 1
                print(f"  [CONTAMINATION] {art_id}: Found {found_patterns}")
                
            # Check for truncation suspicion (e.g., odd number of quotes in Verbatim Evidence)
            # The format is usually: <Reason>. Verbatim Evidence: "<Quote>"
            verb_idx = notes_str.find('Verbatim Evidence:')
            if verb_idx != -1:
                verb_text = notes_str[verb_idx:]
                # Check if it ends abruptly (e.g. missing closing quote)
                if not verb_text.rstrip().endswith('"'):
                    # But it might just not have a closing quote naturally? Let's check odd quotes
                    quotes = verb_text.count('"') - verb_text.count('\\"')
                    if quotes % 2 != 0:
                        truncation_suspicions += 1
                        print(f"  [TRUNCATION?] {art_id}: Verbatim evidence has odd number of quotes. Ends with: {notes_str[-30:]}")
                        
    print(f"\n--- Summary for {filepath.split(chr(92))[-1]} ---")
    print(f"Total AI Processed rows checked: {total_processed}")
    print(f"Contaminations found: {contamination_count}")
    print(f"Truncation suspicions (odd quotes): {truncation_suspicions}")
    print(f"Empty notes (Missing reasoning/evidence): {empty_notes_count}")

for f in FILES_TO_CHECK:
    scan_excel(f)
