import openpyxl

# Liu 2024 Non-BS Variables Mapping
LIU_MEASURES = {
    "mutual trust": {"items": 11, "min": 1, "max": 7, "type": "1 = Self-report", "note": "expatriates and coworkers", "source": "McAllister (1995)"},
    "cognition-based mutual trust": {"items": 11, "min": 1, "max": 7, "type": "1 = Self-report", "note": "expatriates and coworkers", "source": "McAllister (1995)"},
    "affect-based mutual trust": {"items": 11, "min": 1, "max": 7, "type": "1 = Self-report", "note": "expatriates and coworkers", "source": "McAllister (1995)"},
    "role stressors": {"items": 17, "min": 1, "max": 7, "type": "1 = Self-report", "note": "expatriates", "source": "Rizzo et al. (1970); Beehr et al. (1976)"},
    "subsidiary identification": {"items": 6, "min": 1, "max": 7, "type": "1 = Self-report", "note": "expatriates", "source": "Mael & Ashforth (1992)"},
    "mne identification": {"items": 6, "min": 1, "max": 7, "type": "3 = Others (e.g. coworker/peer subordinate customer/client)", "note": "coworkers", "source": "Mael & Ashforth (1992)"},
    "emotional exhaustion": {"items": 9, "min": 1, "max": 7, "type": "1 = Self-report", "note": "expatriates", "source": "Maslach & Jackson (1981)"},
    "outgroup categorization": {"items": 7, "min": 1, "max": 7, "type": "3 = Others (e.g. coworker/peer subordinate customer/client)", "note": "coworkers", "source": "Leonardelli & Toh (2011)"}
}

LIU_BS_ITEMS = {
    "functional boundary-spanning": 9,
    "linguistic boundary-spanning": 7,
    "cultural boundary-spanning": 10,
    "expatriates’ boundary-spanning": 26
}

# Marrone 2007 Non-BS Variables Mapping
MARRONE_MEASURES = {
    "boundary-spanning self-efficacy": {"items": 8, "min": 1, "max": 5, "type": "1 = Self-report", "note": "Self", "source": "Parker (1998)"},
    "asian ethnicity": {"items": 1, "min": 0, "max": 1, "type": "1 = Self-report", "note": "Self", "source": 999},
    "gmat score": {"items": 1, "min": 999, "max": 999, "type": "4 = Objective", "note": "Objective", "source": 999},
    "role overload": {"items": 3, "min": 1, "max": 5, "type": "1 = Self-report", "note": "Self", "source": "Beehr et al. (1976)"}
}

def clean_key(val):
    if not val: return ""
    return str(val).lower().replace("expatriates’", "expatriates'").strip()

if __name__ == "__main__":
    v7_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_Actual Coding Sheet_v7.xlsx"
    v8_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_Actual Coding Sheet_v8.xlsx"
    
    wb = openpyxl.load_workbook(v7_path)
    ws = wb.active
    
    updated_rows = 0
    
    for r in range(4, ws.max_row + 1):
        article_id = ws.cell(row=r, column=2).value
        if not article_id: continue
        
        # Non-BS Variable name is in Col 45
        non_bs_var = ws.cell(row=r, column=45).value
        # BS Variable name is in Col 41
        bs_var = ws.cell(row=r, column=41).value
        
        key = clean_key(non_bs_var)
        bs_key = clean_key(bs_var)
        
        if "BSMA0006" in str(article_id):
            # Update specific BS items if applicable
            for k, items in LIU_BS_ITEMS.items():
                if k in bs_key:
                    ws.cell(row=r, column=27).value = items  # Number of Items
            
            # Map Non-BS variable
            found = False
            for map_key, data in LIU_MEASURES.items():
                if map_key in key:
                    ws.cell(row=r, column=34).value = data["items"]
                    ws.cell(row=r, column=35).value = data["min"]
                    ws.cell(row=r, column=36).value = data["max"]
                    ws.cell(row=r, column=37).value = data["type"]
                    ws.cell(row=r, column=38).value = data["note"]
                    ws.cell(row=r, column=39).value = data["source"]
                    updated_rows += 1
                    found = True
                    break
            
            # If the Non-BS variable is actually another BS variable (which happens in matrices)
            if not found:
                for k, items in LIU_BS_ITEMS.items():
                    if k in key:
                        ws.cell(row=r, column=34).value = items
                        ws.cell(row=r, column=35).value = 1
                        ws.cell(row=r, column=36).value = 7
                        ws.cell(row=r, column=37).value = "3 = Others (e.g. coworker/peer subordinate customer/client)"
                        ws.cell(row=r, column=38).value = "host-country coworkers"
                        ws.cell(row=r, column=39).value = "Liu et al. (2024)"
                        updated_rows += 1
                        break

        elif "BSMA0008" in str(article_id):
            found = False
            for map_key, data in MARRONE_MEASURES.items():
                if map_key in key:
                    ws.cell(row=r, column=34).value = data["items"]
                    ws.cell(row=r, column=35).value = data["min"]
                    ws.cell(row=r, column=36).value = data["max"]
                    ws.cell(row=r, column=37).value = data["type"]
                    ws.cell(row=r, column=38).value = data["note"]
                    ws.cell(row=r, column=39).value = data["source"]
                    updated_rows += 1
                    found = True
                    break
            
            # If Non-BS var is another BS var
            if not found and "boundary-spanning role" in key:
                ws.cell(row=r, column=34).value = 999
                ws.cell(row=r, column=35).value = 1
                ws.cell(row=r, column=36).value = 5
                ws.cell(row=r, column=37).value = "1 = Self-report"
                ws.cell(row=r, column=38).value = "Self"
                ws.cell(row=r, column=39).value = "999"
                updated_rows += 1
                
    wb.save(v8_path)
    print(f"v8 generation complete. Updated {updated_rows} rows with precise measure descriptors!")
