import os
import csv
import glob
import json
import PyPDF2
import openpyxl

def get_next_id(excel_path, csv_path):
    max_id = 0
    # Check Excel
    if os.path.exists(excel_path):
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            ws = wb.active
            for r in range(4, ws.max_row + 1):
                val = ws.cell(row=r, column=2).value
                if val and str(val).startswith("BSMA"):
                    try:
                        num = int(str(val)[4:])
                        if num > max_id:
                            max_id = num
                    except ValueError:
                        pass
        except Exception as e:
            print(f"Error reading Excel for max ID: {e}")

    # Check CSV
    if os.path.exists(csv_path):
        try:
            with open(csv_path, "r", encoding="utf-8") as f:
                content = f.read()
                if content.startswith('\ufeff'):
                    content = content[1:]
                lines = content.splitlines()
                reader = csv.DictReader(lines)
                for row in reader:
                    cleaned_row = {k.strip(): v for k, v in row.items() if k is not None}
                    val = cleaned_row.get("Article_ID")
                    if val and str(val).startswith("BSMA"):
                        try:
                            num = int(str(val)[4:])
                            if num > max_id:
                                max_id = num
                        except ValueError:
                            pass
        except Exception as e:
            print(f"Error reading CSV for max ID: {e}")

    next_num = max_id + 1
    return f"BSMA{next_num:04d}"

def main():
    excel_path = "BSMA_Master_Coding_Sheet.xlsx"
    csv_path = "03_Archives_and_Backups/batch_queue.csv"
    pdf_dir = "01_Academic_Papers"
    txt_dir = "03_Archives_and_Backups/pdf_texts"

    if not os.path.exists(csv_path):
        print(json.dumps({"error": f"Queue file not found at {csv_path}"}))
        return

    # Read queue
    rows = []
    headers = []
    target_row = None
    target_idx = -1

    with open(csv_path, "r", encoding="utf-8") as f:
        # Handle potential UTF-8 BOM
        content = f.read()
        if content.startswith('\ufeff'):
            content = content[1:]
        
        lines = content.splitlines()
        reader = csv.reader(lines)
        headers = [h.strip() for h in next(reader)]
        for idx, r in enumerate(reader):
            row_dict = dict(zip(headers, r))
            rows.append(row_dict)
            if target_row is None and row_dict.get("Status") == "PENDING":
                target_row = row_dict
                target_idx = idx

    if not target_row:
        print(json.dumps({"status": "FINISHED", "message": "No pending papers in the queue."}))
        return

    filename = target_row.get("Filename")
    
    # Locate PDF
    pdf_path = None
    direct_path = os.path.join(pdf_dir, filename)
    if os.path.exists(direct_path):
        pdf_path = direct_path
    else:
        # Search recursively
        for root, _, files in os.walk(pdf_dir):
            if filename in files:
                pdf_path = os.path.join(root, filename)
                break
            for f in files:
                if f.lower() == filename.lower():
                    pdf_path = os.path.join(root, f)
                    break
    
    if not pdf_path:
        # Update status to FAILED in queue
        target_row["Status"] = "FAILED"
        target_row["Last_Updated"] = "File not found"
        write_queue(csv_path, headers, rows)
        print(json.dumps({"error": f"PDF file not found: {filename}", "article_id": "UNKNOWN"}))
        return

    # Allocate ID
    article_id = get_next_id(excel_path, csv_path)
    txt_path = os.path.join(txt_dir, f"{article_id}.txt")

    # Extract text
    try:
        os.makedirs(txt_dir, exist_ok=True)
        with open(pdf_path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            text = ""
            for page in reader.pages:
                text += page.extract_text() + "\n"
        with open(txt_path, "w", encoding="utf-8") as out_f:
            out_f.write(text)
    except Exception as e:
        target_row["Status"] = "FAILED"
        target_row["Last_Updated"] = f"Extraction failed: {str(e)}"
        write_queue(csv_path, headers, rows)
        print(json.dumps({"error": f"Failed to extract text from PDF: {str(e)}", "article_id": article_id}))
        return

    # Update row
    target_row["Article_ID"] = article_id
    target_row["Status"] = "PROCESSING"
    import datetime
    target_row["Last_Updated"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Write back to CSV
    write_queue(csv_path, headers, rows)

    # Print success info
    print(json.dumps({
        "status": "SUCCESS",
        "article_id": article_id,
        "filename": filename,
        "pdf_path": pdf_path,
        "text_path": txt_path
    }))

def write_queue(csv_path, headers, rows):
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for r in rows:
            writer.writerow([r.get(h, "") for h in headers])

if __name__ == "__main__":
    main()
