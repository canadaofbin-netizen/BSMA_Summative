import os
import json
import glob
import openpyxl
import argparse

PDF_DIR = os.path.join(os.getcwd(), "01_Academic_Papers")
PAYLOAD_OUT = os.path.join(os.getcwd(), "scratch", "subagents_payload.json")

def find_pdf_for_id(art_id):
    try:
        num = int(art_id.replace("BSMA", ""))
        prefix = f"[{num}]"
        for file in os.listdir(PDF_DIR):
            if file.startswith(prefix) and file.lower().endswith(".pdf"):
                return os.path.join(PDF_DIR, file)
    except Exception:
        pass
    return None

def prep_swarm(excel_path, batch_size=40):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    pending = []
    for row in range(4, ws.max_row + 1):
        art_id = str(ws.cell(row, 2).value).strip()
        status = ws.cell(row, 5).value
        
        if not status or str(status).strip() == "" or str(status) == "None":
            if art_id and art_id != "None":
                pdf_path = find_pdf_for_id(art_id)
                if pdf_path:
                    # Convert to forward slashes for prompt robustness
                    pdf_path_clean = pdf_path.replace("\\", "/")
                    pending.append((art_id, pdf_path_clean))
                else:
                    print(f"WARNING: Missing PDF file for {art_id}. Skipping.")
                
                if len(pending) >= batch_size:
                    break
                    
    if not pending:
        print("No more pending papers found!")
        return
        
    print(f"Preparing Swarm for {len(pending)} papers: {pending[0][0]} to {pending[-1][0]}")
    
    # Critical Override Preamble (post-Swarm Adjudication + Smoke Test v2)
    preamble = (
        "=== CRITICAL OVERRIDE RULES (CHECK BEFORE ANY EXCLUSION) ===\n\n"
        "1. LEADER BSB OVERRIDE (Rule 16): If the paper measures a leader/manager's "
        "boundary spanning behavior as rated by subordinates -> MUST INCLUDE. "
        "Leader BSB IS BSB. Do NOT exclude under Code 1 or 'cross-entity' logic. "
        "EXCEPTION: If items measure internal demographic diversity (e.g., Blau's index), "
        "that is inclusive leadership, NOT BSB -> Exclude.\n\n"
        "2. INTRA-ORG BSB OVERRIDE (Rule 17): If the paper measures boundary spanning "
        "across internal departments/functions (e.g., IT <-> Business, R&D <-> Marketing, "
        "cross-lab knowledge seeking) -> MUST INCLUDE. Intra-Org BSB IS BSB.\n\n"
        "3. CEO BSB EXCEPTION (Rule 15): If a CEO rates their OWN individual behavior "
        "using 'I' referent items (e.g., 'I solicit information from external channels') "
        "and N=firms=N=CEOs -> Valid individual-level BSB. INCLUDE.\n\n"
        "4. 1:1 MATCHED DYAD SAFE HARBOR: If a study uses 1-to-1 matched pairs "
        "(e.g., 1 expatriate paired with 1 coworker, N=177 unique individuals) "
        "-> Statistical independence is preserved. Do NOT exclude under Code 3. "
        "ONLY exclude under Code 3 if N = relationships (e.g., 87 people -> 673 ties).\n\n"
        "5. PROXY TRAP VERIFICATION (Rule 15): Always check the grammatical subject "
        "of survey items. 'I' = individual-level (valid). 'We/Our team/My organization' "
        "= macro-level key informant (exclude Code 3).\n\n"
        "6. AI TOOL INTERACTION (Rule 13 strict): BSB must be interpersonal (human-to-human). "
        "Using AI tools (ChatGPT) to gather information is NOT BSB -> Exclude Code 1.\n\n"
        "7. NETWORK-BASED BSB INCLUSION (Rule 17+18 supplement): If a study measures "
        "cross-boundary ties using a network generator that asks about PURPOSIVE actions "
        "(e.g., 'who do you go to for information or knowledge on work-related topics'), "
        "the resulting network variables (bridging ties, cross-boundary contacts, degree "
        "centrality) ARE valid BSB operationalizations -> INCLUDE. Do NOT exclude simply "
        "because BSB is measured via social network analysis rather than a Likert scale. "
        "The key test is whether the network generator question implies purposive boundary "
        "spanning action. ONLY Burt's structural constraint is explicitly invalid.\n\n"
        "8. IMPLANTED BOUNDARY SPANNER / KNOWLEDGE EXCHANGE INCLUSION: If a study's "
        "sample consists of employees whose JOB ROLE is boundary spanning (e.g., logistics "
        "implants on-site at client facilities, expatriates) AND the study measures their "
        "knowledge exchange, coordination, or information-sharing behavior with the partner "
        "organization, this IS a valid BSB construct -> INCLUDE. Do NOT exclude just because "
        "the paper labels it 'knowledge exchange' instead of 'boundary spanning behavior'. "
        "The construct captures the same purposive cross-boundary activity.\n\n"
        "=== FEW-SHOT JUDGMENT EXAMPLES ===\n\n"
        "EXAMPLE 1 (INCLUDE - Leader BSB):\n"
        "Paper: 'Leader boundary-spanning behavior and creative behavior'\n"
        "Sample: N=260 employees rating their supervisor's BSB\n"
        "-> Rule 16 OVERRIDE: Leader BSB rated by subordinates = INCLUDE\n\n"
        "EXAMPLE 2 (INCLUDE - Intra-Org BSB via Network):\n"
        "Paper: 'Activating Cross-Boundary Knowledge: The Role of Simmelian Ties'\n"
        "Sample: N=245 R&D scientists spanning across internal research labs\n"
        "Network generator asks 'who do you go to for information or knowledge'\n"
        "-> Rule 7: Purposive network ties across internal boundaries = valid BSB = INCLUDE\n"
        "-> Do NOT exclude because BSB is measured as network ties rather than a Likert scale\n\n"
        "EXAMPLE 3 (INCLUDE - Implanted Boundary Spanner):\n"
        "Paper: 'Enhancing Dyadic Performance Through Boundary Spanners and Innovation'\n"
        "Sample: N=81 logistics implants on-site at customer facilities (1:1 matched dyads)\n"
        "Measured construct: Knowledge Exchange between implant and host firm\n"
        "-> Rule 8: Implanted boundary spanners' knowledge exchange IS BSB = INCLUDE\n"
        "-> Rule 4: 1:1 matched pairs preserve statistical independence\n\n"
        "EXAMPLE 4 (EXCLUDE - Nested/Repeated Data):\n"
        "Paper: 'Boundary Spanning by Public Managers'\n"
        "Sample: N=158 observations, but 98 managers each rated 2 projects\n"
        "-> Code 3: Repeated measures from same individuals violates independence\n\n"
        "EXAMPLE 5 (EXCLUDE - AI Interaction, not BSB):\n"
        "Paper: 'How AI use contributes to employee competitive advantage'\n"
        "Sample: Employees using ChatGPT to 'solicit information from external channels'\n"
        "-> Code 1: Human-computer interaction, not interpersonal BSB\n\n"
        "=== OUTPUT FORMAT (MANDATORY FIELD ORDER) ===\n\n"
        "You MUST output your JSON in EXACTLY this order:\n"
        "{\n"
        '  "BSMA_ID": "...",           // 1st - Always preserved\n'
        '  "Verdict": 0 or 1,          // 2nd - Always preserved\n'
        '  "Exclusion_Code": ...,      // 3rd - Always preserved\n'
        '  "Reasoning": "...",          // 4th - Summary of judgment\n'
        '  "Verbatim_Evidence": "..."   // 5th - LAST. Extract MAXIMUM evidence (Rule 30)\n'
        "}\n\n"
        "IMPORTANT: Write Verbatim_Evidence LAST. Extract MAXIMUM evidence "
        "from ALL sections of the PDF (Abstract, Method, Results, Discussion). "
        "Do NOT truncate or summarize quotes. Use full, complete sentences.\n\n"
        "=== END OF PREAMBLE ==="
    )

    subagents = []
    for art_id, pdf_path in pending:
        prompt = f"Now review `{art_id}`. The PDF is at `{pdf_path}`. Output to `scratch/outputs/{art_id}.json`."

        subagents.append({
            "TypeName": "bsma_reviewer_v4",
            "Role": f"Reviewer for {art_id}",
            "Prompt": prompt
        })
        
    os.makedirs(os.path.dirname(PAYLOAD_OUT), exist_ok=True)
    with open(PAYLOAD_OUT, 'w', encoding='utf-8') as f:
        json.dump(subagents, f, indent=4)
        
    print(f"Payload for {len(pending)} subagents saved to {PAYLOAD_OUT}")
    print(f"-> Antigravity: You can now read this file and pass it to `invoke_subagent`.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--batch", type=int, default=40, help="Number of subagents to prep")
    parser.add_argument("--excel", type=str, required=True, help="Absolute path to the target Excel file")
    args = parser.parse_args()
    prep_swarm(args.excel, args.batch)
