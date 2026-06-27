import re
import openpyxl

marrone_md = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\03_Shadow_Reports\Marrone_2007_Shadow_Report.md"
akaho_md = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\03_Shadow_Reports\Akaho_2024_Shadow_Report.md"
excel_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_Actual Coding Sheet_v3.xlsx"

wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# Find max row
r_idx = 1
while ws.cell(row=r_idx, column=1).value is not None:
    r_idx += 1

def clean(val):
    val = val.strip()
    val = val.replace('**', '').replace('*', '')
    if val == '999': return 999
    if val == '1': return 1
    if val == '0': return 0
    try: return float(val)
    except: return val

# Insert Akaho (BSMA0004) - Excluded
ws.cell(row=r_idx, column=1, value="BSMA0004")
ws.cell(row=r_idx, column=2, value="KY")
ws.cell(row=r_idx, column=3, value="BSMA0004.1")
ws.cell(row=r_idx, column=4, value="BSMA0004.1.1")
ws.cell(row=r_idx, column=5, value=0) # Exclude
ws.cell(row=r_idx, column=8, value="Yuma Akaho")
ws.cell(row=r_idx, column=9, value=2024)
ws.cell(row=r_idx, column=12, value="4 = Non-primary study") # Non-primary
ws.cell(row=r_idx, column=16, value="Excluded based on Shadow Report: Conceptual/theoretical paper.")
r_idx += 1

# Parse Marrone (BSMA0001)
with open(marrone_md, 'r', encoding='utf-8') as f:
    text = f.read()

tables = re.findall(r'\| Effect Size ID.*?(?=\n\n|\Z)', text, re.DOTALL)

marrone_count = 1

for t in tables:
    lines = t.strip().split('\n')
    for line in lines:
        if line.startswith('| Effect Size ID') or line.startswith('|---') or line.startswith('| :---'):
            continue
        if line.startswith('|'):
            cols = [c.strip() for c in line.split("|")[1:-1]]
            if len(cols) < 9: continue
            
            # | Effect Size ID | BS Variable Name | Non-BS Variable Name | Correlation ($r$) | Pairwise/Var $N$ | BS Items | BS Alpha | Non-BS Items | Non-BS Alpha | Study Design |
            corr_var = clean(cols[2])
            r_val = clean(cols[3])
            n_val = clean(cols[4])
            bs_alpha = clean(cols[6])
            non_bs_alpha = clean(cols[8])
            
            eff_id = f"BSMA0001.1.{marrone_count}"
            marrone_count += 1
            
            # Write to excel
            ws.cell(row=r_idx, column=1, value="BSMA0001")
            ws.cell(row=r_idx, column=2, value="KY")
            ws.cell(row=r_idx, column=3, value="BSMA0001.1")
            ws.cell(row=r_idx, column=4, value=eff_id)
            ws.cell(row=r_idx, column=5, value=1)
            ws.cell(row=r_idx, column=8, value="Jennifer A. Marrone, Paul E. Tesluk, Jay B. Carson")
            ws.cell(row=r_idx, column=9, value=2007)
            ws.cell(row=r_idx, column=12, value="1 = Journal article")
            ws.cell(row=r_idx, column=21, value="1 = No")
            ws.cell(row=r_idx, column=22, value=999)
            ws.cell(row=r_idx, column=24, value=999)
            ws.cell(row=r_idx, column=25, value=999)
            ws.cell(row=r_idx, column=26, value="MBA students in consulting")
            ws.cell(row=r_idx, column=27, value="USA")
            ws.cell(row=r_idx, column=28, value=1)
            ws.cell(row=r_idx, column=29, value=5)
            ws.cell(row=r_idx, column=30, value=3)
            ws.cell(row=r_idx, column=32, value="Boundary-spanning behavior. Closely following existing research...")
            
            ws.cell(row=r_idx, column=36, value=corr_var)
            ws.cell(row=r_idx, column=42, value=non_bs_alpha)
            ws.cell(row=r_idx, column=44, value=n_val)
            ws.cell(row=r_idx, column=45, value=r_val)
            
            r_idx += 1

wb.save(excel_path)
print(f"Inserted Akaho (1 row) and Marrone ({marrone_count - 1} rows) into v3.xlsx. Total new rows: {r_idx - 1}")
