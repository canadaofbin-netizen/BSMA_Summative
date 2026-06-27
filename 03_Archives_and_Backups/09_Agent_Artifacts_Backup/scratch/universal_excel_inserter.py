import openpyxl
import re
import os

def clean_val(val):
    if val is None: return None
    val = str(val).strip()
    val = val.replace('\\', '').replace('**', '').replace('*', '')
    if val == '999' or val == '999 [FLAGGED]': return 999
    try: return float(val)
    except ValueError: return val

def split_compound(val, index):
    val = clean_val(val)
    if isinstance(val, str) and ' / ' in val:
        parts = val.split(' / ')
        if len(parts) > index:
            return clean_val(parts[index])
    return val

def parse_markdown_table(text):
    tables = re.findall(r'\| Effect Size ID.*?(?=\n\n|\Z)', text, re.DOTALL)
    parsed_tables = []
    for t in tables:
        lines = t.strip().split('\n')
        if len(lines) < 3: continue
        headers = [c.strip().replace('*', '').lower() for c in lines[0].split('|')[1:-1]]
        rows = []
        for line in lines[2:]:
            if not line.startswith('|'): break
            cols = [c.strip() for c in line.split('|')[1:-1]]
            row_dict = dict(zip(headers, cols))
            rows.append(row_dict)
        parsed_tables.extend(rows)
    return parsed_tables

def append_paper(ws, md_path, article_id_num, is_first_insertion):
    with open(md_path, 'r', encoding='utf-8') as f:
        text = f.read()

    fname = os.path.basename(md_path)
    
    author = "Unknown"
    year = 2024
    if "Liu" in fname:
        author = "Ting Liu, Tomoki Sekiguchi, Jiayin Qin, Ya Xi Shen"
        year = 2024
    elif "Akaho" in fname:
        author = "Yuma Akaho"
        year = 2024
    elif "Marrone" in fname:
        author = "Jennifer A. Marrone, Paul E. Tesluk, Jay B. Carson"
        year = 2007

    if not is_first_insertion:
        # Blank row rule
        ws.append([])

    rows = parse_markdown_table(text)
    
    if len(rows) == 0:
        row = [None]*50
        row[0] = "KY"
        row[1] = f"BSMA{article_id_num:04d}"
        row[2] = f"BSMA{article_id_num:04d}.1"
        row[3] = f"BSMA{article_id_num:04d}.1.1"
        row[4] = "0 = Exclude"
        row[7] = author
        row[8] = year
        row[11] = "4 = Non-primary study"
        ws.append(row)
        return

    sample_id = f"BSMA{article_id_num:04d}.1"
    
    for i, r_dict in enumerate(rows):
        row = [None]*50
        eff_id = f"{sample_id}.{i+1}"
        
        row[0] = "KY"
        row[1] = f"BSMA{article_id_num:04d}"
        row[2] = sample_id
        row[3] = eff_id
        
        inc = 1
        if "included?" in r_dict:
            v = str(r_dict["included?"]).lower()
            if "0" in v or "exclude" in v: inc = 0
        row[4] = "1 = Include" if inc == 1 else "0 = Exclude"
        
        row[7] = author
        row[8] = year
        row[11] = "1 = Journal article"
        row[20] = "1 = No"
        
        row[21] = 999
        row[22] = 999
        row[23] = 999
        row[24] = 999
        
        if "Liu" in fname:
            row[25] = "Expatriate employees in tech firms"
            row[26] = "China"
            row[40] = "Expatriates' boundary-spanning"
            
            row[41] = split_compound(r_dict.get("mean"), 0)
            row[42] = split_compound(r_dict.get("sd"), 0)
            row[43] = split_compound(r_dict.get("alpha"), 0)
            
            row[44] = clean_val(r_dict.get("correlate variable"))
            row[45] = split_compound(r_dict.get("mean"), 1)
            row[46] = split_compound(r_dict.get("sd"), 1)
            row[47] = split_compound(r_dict.get("alpha"), 1)
            
            row[48] = clean_val(r_dict.get("r"))

        elif "Marrone" in fname:
            row[25] = "MBA students in consulting"
            row[26] = "USA"
            row[40] = "Boundary-spanning behavior"
            row[43] = 0.86
            
            row[44] = clean_val(r_dict.get("non-bs variable name"))
            row[47] = clean_val(r_dict.get("non-bs alpha"))
            row[48] = clean_val(r_dict.get("correlation ($r$)"))
            
        ws.append(row)


if __name__ == "__main__":
    v2_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_Actual Coding Sheet_v2.xlsx"
    v5_path = r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\BSMA_Actual Coding Sheet_v5.xlsx"
    
    wb = openpyxl.load_workbook(v2_path)
    ws = wb.active
    
    max_id = 0
    for r in range(4, ws.max_row + 1):
        val = ws.cell(row=r, column=2).value
        if val and str(val).startswith('BSMA'):
            m = re.search(r'\d+', str(val))
            if m: max_id = max(max_id, int(m.group()))
            
    print(f"Base Max Article ID: {max_id}")
    
    reports = [
        r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\03_Shadow_Reports\Liu_2024_Shadow_Report.md",
        r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\03_Shadow_Reports\Akaho_2024_Shadow_Report.md",
        r"g:\My Drive\UCL\BSMA\BSMA ANTIGRAVITY\03_Shadow_Reports\Marrone_2007_Shadow_Report.md"
    ]
    
    is_first = True
    for report in reports:
        max_id += 1
        print(f"Appending {os.path.basename(report)} as BSMA{max_id:04d}...")
        append_paper(ws, report, max_id, is_first)
        is_first = False
        
    wb.save(v5_path)
    print("Generated v5.xlsx successfully!")
